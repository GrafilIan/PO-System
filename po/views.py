import json
import zipfile
from decimal import Decimal

import pandas as pd
import openpyxl
from django.core.exceptions import ValidationError
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from openpyxl.workbook import Workbook

from JubanShop.views import juban_inventory_table
from .forms import PurchaseOrderForm, UploadFileForm, ItemInventoryBulkForm, PurchaseOrderBulkForm, \
    ItemInventoryListForm, ItemInventoryQuantityForm, StockInHistoryForm, EditRemarksForm
from .models import PurchaseOrder, ArchiveFolder, ItemInventory, SupplierFolder, InventoryHistory, SiteInventoryFolder, \
    ClientInventoryFolder, Cart, poCart, ItemCodeList, InventorySupplierFolder, StockInHistory
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Sum


# Create your views here.

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password')
            return render(request, 'authentication/login.html', {'error': 'Invalid credentials'})
    return render(request, 'authentication/login.html')


@login_required
def dashboard_view(request):
    if request.user.groups.filter(name='Accountant').exists():
        return render(request, 'dashboards/accountant_dashboard.html')
    elif request.user.groups.filter(name='Front Desk').exists():
        return purchase_order_list(request)
    elif request.user.groups.filter(name='Inventory Manager').exists():
        return inventory_table(request)
    elif request.user.groups.filter(name='Juban Inventory Manager').exists():
        return juban_inventory_table(request)  # Route to the Juban-specific inventory view
    elif request.user.is_superuser:
        return render(request, 'dashboards/front_desk_dashboard.html')
    else:
        return HttpResponse('No role assigned')


def logout_view(request):
    logout(request)
    return redirect('login')


# For Creating or Adding Records

def purchase_order_create(request):
    if request.method == 'POST':
        form = PurchaseOrderForm(request.POST)
        if form.is_valid():
            # Save the new purchase order
            purchase_order = form.save(commit=False)

            # Get the supplier name from the purchase order
            supplier_name = purchase_order.supplier

            if supplier_name:
                # Create or get the SupplierFolder associated with this supplier
                folder, created = SupplierFolder.objects.get_or_create(name=supplier_name)

                # Associate the new purchase order with the supplier's folder
                purchase_order.supplier_folder = folder
                purchase_order.save()

                # Ensure all existing purchase orders with the same supplier are associated with the folder
                matching_orders = PurchaseOrder.objects.filter(supplier=supplier_name, supplier_folder__isnull=True)
                for order in matching_orders:
                    order.supplier_folder = folder
                    order.save()

            purchase_order.site_delivered = purchase_order.site_delivered.strip().upper()
            # Removed the following code related to ItemInventory:
            #
            # if purchase_order.site_delivered in ['BTCS', 'BTCS WH']:
            #     # Check if an inventory record for the item exists
            #     inventory_item = ItemInventory.objects.filter(
            #         supplier=purchase_order.supplier,
            #         po_product_name=purchase_order.particulars,
            #         unit=purchase_order.unit,
            #         price=purchase_order.price,
            #     ).first()
            #
            #     if inventory_item:
            #         # Update existing inventory record
            #         inventory_item.quantity_in += purchase_order.quantity
            #     else:
            #         # Create a new inventory record
            #         inventory_item = ItemInventory(
            #             supplier=purchase_order.supplier,
            #             po_product_name=purchase_order.particulars,
            #             unit=purchase_order.unit,
            #             quantity_in=purchase_order.quantity,
            #             price=purchase_order.price,
            #             delivery_ref=purchase_order.delivery_ref,
            #             delivery_no=purchase_order.delivery_no,
            #             invoice_type=purchase_order.invoice_type,
            #             invoice_no=purchase_order.invoice_no,
            #         )
            #     inventory_item.save()

            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'status': 'error'})
    else:
        form = PurchaseOrderForm()

    return render(request, 'records/purchase_order_form.html', {'form': form})


def purchase_order_edit(request, id):
    order = get_object_or_404(PurchaseOrder, id=id)

    if request.method == 'POST':
        form = PurchaseOrderForm(request.POST, instance=order)
        if form.is_valid():
            purchase_order = form.save()

            purchase_order.site_delivered = purchase_order.site_delivered.strip().upper()
            # Removed the following code related to ItemInventory:
            #
            # if purchase_order.site_delivered in ['BTCS', 'BTCS WH']:
            #     # Update or create inventory item
            #     inventory_item = ItemInventory.objects.filter(
            #         supplier=purchase_order.supplier,
            #         po_product_name=purchase_order.particulars,
            #         unit=purchase_order.unit,
            #     ).first()
            #
            #     if inventory_item:
            #         # Update existing inventory record
            #         inventory_item.quantity_in += purchase_order.quantity
            #     else:
            #         # Create a new inventory record
            #         inventory_item = ItemInventory(
            #             supplier=purchase_order.supplier,
            #             po_product_name=purchase_order.particulars,
            #             unit=purchase_order.unit,
            #             quantity_in=purchase_order.quantity,
            #             price=purchase_order.price,
            #         )
            #
            #     # Save the inventory item
            #     inventory_item.save()

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success'})
            return redirect('dashboard')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error'})
    else:
        form = PurchaseOrderForm(instance=order)

    return render(request, 'records/purchase_order_edit.html', {'form': form, 'order': order})


def purchase_order_list(request):
    query = request.GET.get('q')  # Single search input
    page_number = request.GET.get('page', 1)

    # Dictionary to map month names to numbers
    month_mapping = {
        "January": "1", "February": "2", "March": "3", "April": "4",
        "May": "5", "June": "6", "July": "7", "August": "8",
        "September": "9", "October": "10", "November": "11", "December": "12"
    }

    # Initial queryset, excluding archived or foldered orders
    orders_list = PurchaseOrder.objects.filter(folder__isnull=True, archived=False)

    # If there's a search query, filter the orders accordingly
    if query:
        try:
            # Try to interpret the query as a month number (e.g., "08" for August)
            month_number = int(query)
            orders_list = orders_list.filter(date__month=month_number)
        except ValueError:
            try:
                # Try to interpret the query as a full month name only (e.g., "August")
                month_number = month_mapping.get(query.strip().capitalize())
                if month_number:
                    orders_list = orders_list.filter(date__month=month_number)
                else:
                    raise ValueError("Not a valid month name")
            except ValueError:
                # If not a date, treat it as a general text search
                orders_list = orders_list.filter(
                    Q(po_number__icontains=query) |
                    Q(purchaser__icontains=query) |
                    Q(brand__icontains=query) |
                    Q(item_code__icontains=query) |
                    Q(particulars__icontains=query) |
                    Q(quantity__icontains=query) |
                    Q(unit__icontains=query) |
                    Q(price__icontains=query) |
                    Q(total_amount__icontains=query) |
                    Q(site_delivered__icontains=query) |
                    Q(fbbd_ref_number__icontains=query) |
                    Q(remarks__icontains=query) |
                    Q(supplier__icontains=query) |
                    Q(delivery_ref__icontains=query) |
                    Q(delivery_no__icontains=query) |
                    Q(invoice_type__icontains=query) |
                    Q(invoice_no__icontains=query) |
                    Q(payment_req_ref__icontains=query) |
                    Q(payment_details__icontains=query) |
                    Q(remarks2__icontains=query)
                )

    # Pagination
    paginator = Paginator(orders_list, 100)  # Show 20 orders per page
    try:
        orders = paginator.page(page_number)
    except PageNotAnInteger:
        orders = paginator.page(1)
    except EmptyPage:
        orders = paginator.page(paginator.num_pages)

    # Get all folders for the folder dropdown
    folders = ArchiveFolder.objects.all()

    # Render the template with the context
    return render(request, 'dashboards/front_desk_dashboard.html', {
        'orders': orders,
        'query': query,
        'page_number': page_number,
        'folders': folders,
    })


# --------------------------------------For Exporting Records----------------------------------------------
def export_orders_to_excel(request):
    query = request.GET.get('q')  # Search query parameter
    date_query = request.GET.get('date')  # Date query parameter

    # Start with all orders
    orders_list = PurchaseOrder.objects.filter(folder__isnull=True)

    # Apply search filter
    if query:
        orders_list = orders_list.filter(
            Q(date__icontains=query) |
            Q(po_number__icontains=query) |
            Q(purchaser__icontains=query) |
            Q(brand__icontains=query) |
            Q(item_code__icontains=query) |
            Q(particulars__icontains=query) |
            Q(quantity__icontains=query) |
            Q(unit__icontains=query) |
            Q(price__icontains=query) |
            Q(total_amount__icontains=query) |
            Q(site_delivered__icontains=query) |
            Q(fbbd_ref_number__icontains=query) |
            Q(remarks__icontains=query) |
            Q(supplier__icontains=query) |
            Q(delivery_ref__icontains=query) |
            Q(delivery_no__icontains=query) |
            Q(invoice_type__icontains=query) |
            Q(invoice_no__icontains=query) |
            Q(payment_req_ref__icontains=query) |
            Q(payment_details__icontains=query) |
            Q(remarks2__icontains=query)
        )

    # Apply date filter
    if date_query:
        try:
            date_obj = datetime.strptime(date_query, '%b %d, %Y').date()
            orders_list = orders_list.filter(date=date_obj)
        except ValueError:
            try:
                date_obj = datetime.strptime(date_query, '%b %Y')
                orders_list = orders_list.filter(date__year=date_obj.year, date__month=date_obj.month)
            except ValueError:
                try:
                    date_obj = datetime.strptime(date_query, '%B %Y')
                    orders_list = orders_list.filter(date__year=date_obj.year, date__month=date_obj.month)
                except ValueError:
                    try:
                        month_number = int(date_query)
                        orders_list = orders_list.filter(date__month=month_number)
                    except ValueError:
                        try:
                            date_obj = datetime.strptime(date_query, '%B')
                            orders_list = orders_list.filter(date__month=date_obj.month)
                        except ValueError:
                            pass

    # Create a workbook and a sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Purchase Orders'

    header_font = Font(bold=True)
    blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    currency_format = '#,##0.00'

    # Define the headers
    headers = [
        'Date', 'PO Number', 'Purchaser', 'Brand', 'Item Code', 'Particulars',
        'Quantity', 'Unit', 'Price', 'Total Amount',
        'Site Delivered', 'FBBD Ref#', 'Remarks', 'Supplier', 'Delivery Ref#',
        'Delivery No.', 'Invoice Type', 'Invoice No.', 'Payment Req Ref#',
        'Payment Details', 'Remarks2'
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = blue_fill
        cell.value = cell.value.upper() if cell.value is not None else cell.value

    # Populate the sheet with data
    for order in orders_list:
        sheet.append([
            order.date.strftime('%Y-%m-%d') if order.date else 'N/A', order.po_number, order.purchaser, order.brand,
            order.item_code,
            order.particulars, order.quantity, order.unit, order.price, order.total_amount, order.site_delivered,
            order.fbbd_ref_number, order.remarks, order.supplier, order.delivery_ref, order.delivery_no,
            order.invoice_type, order.invoice_no, order.payment_req_ref, order.payment_details, order.remarks2
        ])

    for cell in sheet['I']:  # Assuming price is in column I (index 9)
        if cell.row > 1:  # Skip header row
            cell.number_format = currency_format

    for cell in sheet['J']:  # Assuming total_amount is in column J (index 10)
        if cell.row > 1:  # Skip header row
            cell.number_format = currency_format

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Add extra space for better visibility
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Create an in-memory buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Set the response to return the Excel file
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=PurchaseOrders.xlsx'
    return response


def export_archived_orders_to_excel(request, folder_id):
    try:
        # Get the folder by ID
        folder = ArchiveFolder.objects.get(id=folder_id)
        orders_list = PurchaseOrder.objects.filter(folder=folder)

        # Create a workbook and a sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Archived Orders'

        # Define header styles
        header_font = Font(bold=True)
        blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        currency_format = '#,##0.00'

        # Define the headers
        headers = [
            'Date', 'PO Number', 'Purchaser', 'Brand', 'Item Code', 'Particulars',
            'Quantity', 'Unit', 'Price', 'Total Amount',
            'Site Delivered', 'FBBD Ref#', 'Remarks', 'Supplier', 'Delivery Ref#',
            'Delivery No.', 'Invoice Type', 'Invoice No.', 'Payment Req Ref#',
            'Payment Details', 'Remarks2'
        ]
        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = blue_fill
            cell.value = cell.value.upper() if cell.value is not None else cell.value

        # Populate the sheet with data
        for order in orders_list:
            sheet.append([
                order.date.strftime('%Y-%m-%d') if order.date else 'N/A', order.po_number, order.purchaser, order.brand,
                order.item_code,
                order.particulars, order.quantity, order.unit, order.price, order.total_amount, order.site_delivered,
                order.fbbd_ref_number, order.remarks, order.supplier, order.delivery_ref, order.delivery_no,
                order.invoice_type, order.invoice_no, order.payment_req_ref, order.payment_details, order.remarks2
            ])

        for cell in sheet['I']:  # Assuming price is in column I (index 9)
            if cell.row > 1:  # Skip header row
                cell.number_format = currency_format

        for cell in sheet['J']:  # Assuming total_amount is in column J (index 10)
            if cell.row > 1:  # Skip header row
                cell.number_format = currency_format

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Add extra space for better visibility
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Create an in-memory buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Set the response to return the Excel file
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=ArchivedOrders_{folder.name}.xlsx'
        return response

    except ArchiveFolder.DoesNotExist:
        return HttpResponse("Folder not found", status=404)


def export_supplier_contents(request, folder_id):
    try:
        # Get the supplier folder by ID
        folder = SupplierFolder.objects.get(id=folder_id)
        orders_list = PurchaseOrder.objects.filter(supplier=folder.name)

        # Create a workbook and a sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Supplier Orders'

        # Define header styles
        header_font = Font(bold=True)
        blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        currency_format = '#,##0.00'

        # Define the headers
        headers = [
            'Date', 'PO Number', 'Purchaser', 'Brand', 'Item Code', 'Particulars',
            'Quantity', 'Unit', 'Price', 'Total Amount',
            'Site Delivered', 'FBBD Ref#', 'Remarks', 'Supplier', 'Delivery Ref#',
            'Delivery No.', 'Invoice Type', 'Invoice No.', 'Payment Req Ref#',
            'Payment Details', 'Remarks2'
        ]
        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = blue_fill
            cell.value = cell.value.upper() if cell.value is not None else cell.value

        # Populate the sheet with data
        for order in orders_list:
            sheet.append([
                order.date.strftime('%Y-%m-%d') if order.date else 'N/A', order.po_number, order.purchaser, order.brand,
                order.item_code,
                order.particulars, order.quantity, order.unit, order.price, order.total_amount, order.site_delivered,
                order.fbbd_ref_number, order.remarks, order.supplier, order.delivery_ref, order.delivery_no,
                order.invoice_type, order.invoice_no, order.payment_req_ref, order.payment_details, order.remarks2
            ])

        for cell in sheet['I']:  # Assuming price is in column I (index 9)
            if cell.row > 1:  # Skip header row
                cell.number_format = currency_format

        for cell in sheet['J']:  # Assuming total_amount is in column J (index 10)
            if cell.row > 1:  # Skip header row
                cell.number_format = currency_format

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Add extra space for better visibility
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Create an in-memory buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Set the response to return the Excel file
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=SupplierOrders_{folder.name}.xlsx'
        return response

    except SupplierFolder.DoesNotExist:
        return HttpResponse("Supplier Folder not found", status=404)


def export_all_supplier_folders(request):
    # Create an in-memory buffer to hold the zip file
    buffer = BytesIO()

    # Create a zip file in the buffer
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Iterate through each supplier folder
        for folder in SupplierFolder.objects.all():
            # Get purchase orders for the folder
            orders_list = PurchaseOrder.objects.filter(supplier_folder=folder)

            # Create a workbook and a sheet
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'Supplier Orders'

            # Define header styles
            header_font = Font(bold=True)
            blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
            currency_format = '#,##0.00'

            # Define the headers
            headers = [
                'Date', 'PO Number', 'Purchaser', 'Brand', 'Item Code', 'Particulars',
                'Quantity', 'Unit', 'Price', 'Total Amount',
                'Site Delivered', 'FBBD Ref#', 'Remarks', 'Supplier', 'Delivery Ref#',
                'Delivery No.', 'Invoice Type', 'Invoice No.', 'Payment Req Ref#',
                'Payment Details', 'Remarks2'
            ]
            sheet.append(headers)

            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = blue_fill
                cell.value = cell.value.upper() if cell.value is not None else cell.value

            # Populate the sheet with data
            for order in orders_list:
                sheet.append([
                    order.date.strftime('%Y-%m-%d') if order.date else 'N/A', order.po_number, order.purchaser,
                    order.brand,
                    order.item_code,
                    order.particulars, order.quantity, order.unit, order.price, order.total_amount,
                    order.site_delivered,
                    order.fbbd_ref_number, order.remarks, order.supplier, order.delivery_ref, order.delivery_no,
                    order.invoice_type, order.invoice_no, order.payment_req_ref, order.payment_details, order.remarks2
                ])

            for cell in sheet['I']:  # Assuming price is in column I (index 9)
                if cell.row > 1:  # Skip header row
                    cell.number_format = currency_format

            for cell in sheet['J']:  # Assuming total_amount is in column J (index 10)
                if cell.row > 1:  # Skip header row
                    cell.number_format = currency_format

            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)  # Add extra space for better visibility
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook to a bytes buffer
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            # Add the workbook to the zip file
            zf.writestr(f'SupplierOrders_{folder.name}.xlsx', excel_buffer.getvalue())

    buffer.seek(0)

    # Set the response to return the zip file
    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=supplier_orders.zip'
    return response


def export_transaction_history_to_excel(request):
    query = request.GET.get('q')  # Search query parameter
    date_query = request.GET.get('date')  # Date query parameter

    # Start with all transactions
    transactions = InventoryHistory.objects.exclude(date__isnull=True).order_by('-date')

    # Apply search filter
    if query:
        transactions = transactions.filter(
            Q(date__icontains=query) |
            Q(item_code__icontains=query) |
            Q(supplier__icontains=query) |
            Q(po_product_name__icontains=query) |
            Q(new_product_name__icontains=query) |
            Q(unit__icontains=query) |
            Q(quantity_out__icontains=query) |
            Q(price__icontains=query) |
            Q(total_amount__icontains=query) |
            Q(site_delivered__icontains=query) |
            Q(client__icontains=query) |
            Q(delivery_ref__icontains=query) |
            Q(delivery_no__icontains=query) |
            Q(invoice_type__icontains=query) |
            Q(invoice_no__icontains=query)
        )

    # Apply date filter
    if date_query:
        try:
            date_obj = datetime.strptime(date_query, '%b %d, %Y').date()
            transactions = transactions.filter(date=date_obj)
        except ValueError:
            try:
                date_obj = datetime.strptime(date_query, '%b %Y')
                transactions = transactions.filter(date__year=date_obj.year, date__month=date_obj.month)
            except ValueError:
                try:
                    date_obj = datetime.strptime(date_query, '%B %Y')
                    transactions = transactions.filter(date__year=date_obj.year, date__month=date_obj.month)
                except ValueError:
                    try:
                        month_number = int(date_query)
                        transactions = transactions.filter(date__month=month_number)
                    except ValueError:
                        try:
                            date_obj = datetime.strptime(date_query, '%B')
                            transactions = transactions.filter(date__month=date_obj.month)
                        except ValueError:
                            pass

    # Create a workbook and a sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Transaction History'

    header_font = Font(color='FFFFFF', bold=True)
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    currency_format = '#,##0.00'

    # Define the headers
    headers = [
        'Date', 'Item Code', 'Supplier', 'PO Product Name', 'New Product Name',
        'Unit', 'Quantity Out', 'Price', 'Total Amount', 'Site Delivered',
        'Client', 'Delivery Ref#', 'Delivery No.', 'Invoice Type', 'Invoice#'
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = black_fill
        cell.value = cell.value.upper() if cell.value is not None else cell.value

    # Populate the sheet with data
    for transaction in transactions:
        sheet.append([
            transaction.date.strftime('%Y-%m-%d') if transaction.date else 'N/A',
            transaction.item_code, transaction.supplier, transaction.po_product_name,
            transaction.new_product_name, transaction.unit, transaction.quantity_out,
            transaction.price, transaction.total_amount, transaction.site_delivered,
            transaction.client, transaction.delivery_ref, transaction.delivery_no,
            transaction.invoice_type, transaction.invoice_no
        ])

    # Apply currency format to Price and Total Amount columns
    for cell in sheet['H']:  # Assuming price is in column H (index 8)
        if cell.row > 1:  # Skip header row
            cell.number_format = currency_format

    for cell in sheet['I']:  # Assuming total_amount is in column I (index 9)
        if cell.row > 1:  # Skip header row
            cell.number_format = currency_format

    # Auto-size columns for better visibility
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Create an in-memory buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Set the response to return the Excel file
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=TransactionHistory.xlsx'
    return response


# For Archiving Methods
def create_folder(request):
    if request.method == 'POST':
        folder_name = request.POST.get('folder_name')
        if folder_name:
            try:
                ArchiveFolder.objects.create(name=folder_name)
                return JsonResponse({'status': 'success', 'message': 'Folder created successfully!'})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'Error creating folder: {e}'})
    return render(request, 'archive/list_folders.html')


def list_folders(request):
    folders = ArchiveFolder.objects.all()
    return render(request, 'archive/list_folders.html', {'folders': folders})


def delete_folder(request, folder_id):
    folder = ArchiveFolder.objects.get(id=folder_id)
    orders = PurchaseOrder.objects.filter(folder=folder)

    # Archive the orders instead of deleting them
    orders.update(archived=True)

    # Optionally, you might want to disassociate the orders from the folder
    # orders.update(folder=None)

    folder.delete()
    return redirect('purchase_order_list')


def archive_orders(request, folder_id):
    folder = ArchiveFolder.objects.get(id=folder_id)
    orders = PurchaseOrder.objects.filter(folder=folder)

    paginator = Paginator(orders, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'archive/archive_orders.html', {
        'folder': folder,
        'orders': page_obj
    })


def move_orders_to_folder(request):
    if request.method == 'POST':
        folder_id = request.POST.get('folder')
        order_ids = request.POST.getlist('orders')

        if not folder_id or not order_ids:
            messages.error(request, 'Please select both a folder and orders to move.')
            return redirect('purchase_order_list')

        folder = get_object_or_404(ArchiveFolder, id=folder_id)

        try:
            updated_orders = PurchaseOrder.objects.filter(id__in=order_ids).update(folder=folder)
            if updated_orders > 0:
                messages.success(request, f'{updated_orders} orders successfully moved to {folder.name}.')
            else:
                messages.warning(request, 'No orders were moved.')
        except Exception as e:
            messages.error(request, f'An error occurred while moving orders: {str(e)}')
            return redirect('purchase_order_list')

        return redirect('purchase_order_list')

    return redirect('purchase_order_list')


# For Uploading Files
def handle_uploaded_file(f):
    df = pd.read_excel(f, engine='openpyxl')
    df = df.fillna('')  # Replace NaN with an empty string

    for _, row in df.iterrows():
        # Create or update the PurchaseOrder
        purchase_order = PurchaseOrder(
            date=row.get('DATE'),
            po_number=row.get('PO NUMBER'),
            purchaser=row.get('PURCHASER'),
            brand=row.get('BRAND'),
            item_code=row.get('ITEM CODE'),
            particulars=row.get('PARTICULAR'),
            quantity=row.get('QTY'),
            unit=row.get('UNIT'),
            price=row.get('PRICE'),
            total_amount=row.get('T. AMOUNT'),
            site_delivered=row.get('SITE DELIVERED'),
            fbbd_ref_number=row.get('FBBD REF#'),
            remarks=row.get('REMARKS'),
            supplier=row.get('SUPPLIER'),
            delivery_ref=row.get('DELIVERY REF#'),
            delivery_no=row.get('DELIVERY NO.'),
            invoice_type=row.get('INVOICE TYPE'),
            invoice_no=row.get('INVOICE NO.'),
            payment_req_ref=row.get('PAYMENT REQ REF#'),
            payment_details=row.get('PAYMENT DETAILS'),
            remarks2=row.get('REMARKS2')
        )

        # Check if the supplier folder exists or create a new one
        supplier_name = purchase_order.supplier
        if supplier_name:
            folder, created = SupplierFolder.objects.get_or_create(name=supplier_name)
            # Optionally handle updates to the folder if needed

            # Update the purchase order to reference the SupplierFolder
            purchase_order.supplier_folder = folder

        # Save the purchase order
        purchase_order.save()

    # Removed the following code related to ItemInventory:
    #
    # purchase_order.site_delivered = purchase_order.site_delivered.strip().upper()
    # if purchase_order.site_delivered in ['BTCS', 'BTCS WH']:
    #     # Check if an inventory record for the item exists
    #     inventory_item = ItemInventory.objects.filter(
    #         supplier=purchase_order.supplier,
    #         po_product_name=purchase_order.particulars,
    #         unit=purchase_order.unit,
    #         price=purchase_order.price,
    #     ).first()
    #
    #     if inventory_item:
    #         # Update existing inventory record
    #         inventory_item.quantity_in += purchase_order.quantity
    #     else:
    #         # Create a new inventory record
    #         inventory_item = ItemInventory(
    #             supplier=purchase_order.supplier,
    #             po_product_name=purchase_order.particulars,
    #             unit=purchase_order.unit,
    #             quantity_in=purchase_order.quantity,
    #             price=purchase_order.price,
    #             delivery_ref=purchase_order.delivery_ref,
    #             delivery_no=purchase_order.delivery_no,
    #             invoice_type=purchase_order.invoice_type,
    #             invoice_no=purchase_order.invoice_no,
    #         )
    #
    #     # Save the inventory item
    #     inventory_item.save()


def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                handle_uploaded_file(request.FILES['file'])
                return JsonResponse({'status': 'success', 'message': 'File uploaded and data imported successfully.'})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})
    else:
        form = UploadFileForm()
    return render(request, 'records/upload.html', {'form': form})


# -----------------------------For INVENTORY----------------------------------------------
def inventory_form(request):
    if request.method == 'POST':
        form = ItemInventoryListForm(request.POST)
        po_product_name = request.POST.get('po_product_name')

        # Check if the po_product_name already exists in the database
        if ItemInventory.objects.filter(po_product_name=po_product_name).exists():
            messages.error(request, 'PO Product Name already exists. Please use a different name.')
        elif form.is_valid():
            # Save the form to create the item
            item = form.save(commit=False)  # Get the instance but don't save it to the database yet

            item.quantity_in = item.quantity_in or 0
            item.quantity_out = item.quantity_out or 0

            # Manually calculate the stock
            item.stock = item.quantity_in - item.quantity_out  # Calculate stock based on quantity_in and quantity_out

            # Now save the item with the updated stock value
            item.save()

            messages.success(request, 'Item successfully added!')
            return redirect('inventory_form')  # Redirect to the same form page
        else:
            messages.error(request, 'There was an error adding the item. Please try again.')
    else:
        form = ItemInventoryListForm()

    return render(request, 'inventory/inventory_form.html', {'form': form})


def inventory_table(request):
    query = request.GET.get('q')

    inventory_items = ItemInventory.objects.all().order_by('po_product_name')

    if query:
        inventory_items = inventory_items.filter(
            Q(item_code__icontains=query) |
            Q(supplier__icontains=query) |
            Q(po_product_name__icontains=query) |
            Q(unit__icontains=query) |
            Q(quantity_in__icontains=query) |
            Q(quantity_out__icontains=query) |
            Q(stock__icontains=query)
        )

    for item in inventory_items:
        if not ItemCodeList.objects.filter(item_code=item.item_code, po_product_name=item.po_product_name).exists():
            ItemCodeList.objects.create(
                item_code=item.item_code,
                po_product_name=item.po_product_name,
                unit=item.unit
            )

    paginator = Paginator(inventory_items, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'dashboards/inventory_clerk_dashboard.html', {
        'page_obj': page_obj
    })


def new_records_view(request):
    # Define the timeframe for what constitutes a "new" record (e.g., last 24 hours)
    timeframe = timezone.now() - timezone.timedelta(hours=24)

    # Get all records created in the last 24 hours, ordered by most recent
    new_records = ItemInventory.objects.filter(created_at__gte=timeframe).order_by('-created_at')

    return render(request, 'Inventory/new_records.html', {
        'new_records': new_records
    })


def inventory_edit(request, id):
    inventory_item = get_object_or_404(ItemInventory, id=id)

    if request.method == 'POST':
        form = ItemInventoryQuantityForm(request.POST, instance=inventory_item)
        if form.is_valid():
            updated_item = form.save(commit=False)

            # Check if an existing record with the same supplier and po_product_name already exists
            existing_inventory_item = ItemInventory.objects.filter(
                supplier=updated_item.supplier,
                po_product_name=updated_item.po_product_name
            ).exclude(id=id).first()

            if existing_inventory_item:
                # If a matching record exists, update its quantities and stock
                existing_inventory_item.quantity_in += updated_item.quantity_in
                existing_inventory_item.quantity_out += updated_item.quantity_out
                existing_inventory_item.stock = existing_inventory_item.quantity_in - existing_inventory_item.quantity_out
                existing_inventory_item.save()

                # Optionally, delete the current record or return success message
                inventory_item.delete()  # Remove the current item as it merges with the existing one.
            else:
                # If no matching record, update the current inventory item
                updated_item.stock = updated_item.quantity_in - updated_item.quantity_out
                updated_item.save()

            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'status': 'error', 'errors': form.errors})
    else:
        form = ItemInventoryQuantityForm(instance=inventory_item)

    return render(request, 'inventory/inventory_edit.html', {'form': form, 'item': inventory_item})



def item_code_list(request):
    query = request.GET.get('q')  # Get the search query from the URL
    if query:
        # Filter results based on the search query
        saved_items = ItemCodeList.objects.filter(
            Q(item_code__icontains=query) |
            Q(po_product_name__icontains=query) |
            Q(unit__icontains=query)
        )
    else:
        saved_items = ItemCodeList.objects.all()  # If no search, return all items

    return render(request, 'inventory/item_code_list.html', {'saved_items': saved_items, 'query': query})



def export_inventory_to_excel(request):
    query = request.GET.get('q')  # Search query parameter

    # Start with all inventory items
    inventory_list = ItemInventory.objects.all()

    # Apply search filter
    if query:
        inventory_list = inventory_list.filter(
            Q(item_code__icontains=query) |
            Q(supplier__icontains=query) |
            Q(po_product_name__icontains=query) |
            Q(new_product_name__icontains=query) |
            Q(unit__icontains=query) |
            Q(quantity_in__icontains=query) |
            Q(quantity_out__icontains=query) |
            Q(stock__icontains=query) |
            Q(price__icontains=query) |
            Q(total_amount__icontains=query)
        )

    # Create a workbook and a sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Inventory'

    header_font = Font(bold=True)
    blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    currency_format = '#,##0.00'

    # Define the headers
    headers = [
        'Item Code', 'Supplier', 'Particular', 'New Product Name', 'Unit',
        'Quantity In', 'Quantity Out', 'Stock', 'Price', 'Total Amount'
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = blue_fill

    # Populate the sheet with data
    for item in inventory_list:
        sheet.append([
            item.item_code, item.supplier, item.po_product_name, item.new_product_name,
            item.unit, item.quantity_in, item.quantity_out, item.stock,
            item.price, item.total_amount
        ])

    for cell in sheet['I']:  # Assuming price is in column I (index 9)
        if cell.row > 1:  # Skip header row
            cell.number_format = currency_format

    for cell in sheet['J']:  # Assuming total_amount is in column J (index 10)
        if cell.row > 1:  # Skip header row
            cell.number_format = currency_format

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Add extra space for better visibility
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Create an in-memory buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Set the response to return the Excel file
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Inventory.xlsx'
    return response


def export_site_folder_contents(request, folder_id):
    try:
        # Get the site inventory folder by ID
        folder = SiteInventoryFolder.objects.get(id=folder_id)
        transactions = InventoryHistory.objects.filter(site_inventory_folder=folder)

        # Create a workbook and a sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Site Inventory Records'

        # Define header styles
        header_font = Font(bold=True)
        blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        currency_format = '#,##0.00'

        # Define the headers
        headers = [
            'Transaction ID', 'Site Delivered', 'Date', 'Item', 'Quantity In', 'Quantity Out',
            'Price', 'Total Amount', 'Delivery Ref', 'Delivery No.'
        ]
        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = blue_fill
            cell.value = cell.value.upper() if cell.value is not None else cell.value

        # Populate the sheet with data
        for transaction in transactions:
            sheet.append([
                transaction.id,
                transaction.site_delivered,
                transaction.date.strftime('%Y-%m-%d') if transaction.date else 'N/A',
                str(transaction.item),  # Convert ItemInventory instance to string
                transaction.quantity_in,
                transaction.quantity_out,
                transaction.price,
                transaction.total_amount,
                transaction.delivery_ref,
                transaction.delivery_no,
            ])

        # Apply currency format to price and total_amount columns
        for cell in sheet['G']:  # Assuming price is in column G (index 7)
            if cell.row > 1:  # Skip header row
                cell.number_format = currency_format

        for cell in sheet['H']:  # Assuming total_amount is in column H (index 8)
            if cell.row > 1:  # Skip header row
                cell.number_format = currency_format

        # Adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Add extra space for better visibility
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Create an in-memory buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Set the response to return the Excel file
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=SiteInventory_{folder.name}.xlsx'
        return response

    except SiteInventoryFolder.DoesNotExist:
        return HttpResponse("Site Inventory Folder not found", status=404)


def export_all_site_inventory_folders(request):
    # Create an in-memory buffer to hold the zip file
    buffer = BytesIO()

    # Create a zip file in the buffer
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Iterate through each site inventory folder
        for folder in SiteInventoryFolder.objects.all():
            # Get transactions for the folder
            transactions = InventoryHistory.objects.filter(site_inventory_folder=folder)

            # Create a workbook and a sheet
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'Site Inventory Records'

            # Define header styles
            header_font = Font(bold=True)
            blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
            currency_format = '#,##0.00'

            # Define the headers
            headers = [
                'Transaction ID', 'Site Delivered', 'Date', 'Item', 'Quantity In', 'Quantity Out',
                'Price', 'Total Amount', 'Delivery Ref', 'Delivery No.'
            ]
            sheet.append(headers)

            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = blue_fill
                cell.value = cell.value.upper() if cell.value is not None else cell.value

            # Populate the sheet with data
            for transaction in transactions:
                sheet.append([
                    transaction.id,
                    transaction.site_delivered,
                    transaction.date.strftime('%Y-%m-%d') if transaction.date else 'N/A',
                    str(transaction.item),  # Convert ItemInventory instance to string
                    transaction.quantity_in,
                    transaction.quantity_out,
                    transaction.price,
                    transaction.total_amount,
                    transaction.delivery_ref,
                    transaction.delivery_no,
                ])

            # Apply currency format to price and total_amount columns
            for cell in sheet['G']:  # Assuming price is in column G (index 7)
                if cell.row > 1:  # Skip header row
                    cell.number_format = currency_format

            for cell in sheet['H']:  # Assuming total_amount is in column H (index 8)
                if cell.row > 1:  # Skip header row
                    cell.number_format = currency_format

            # Adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)  # Add extra space for better visibility
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook to a bytes buffer
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            # Add the workbook to the zip file
            zf.writestr(f'SiteInventoryRecords_{folder.name}.xlsx', excel_buffer.getvalue())

    buffer.seek(0)

    # Set the response to return the zip file
    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=site_inventory_folders.zip'
    return response


def export_client_folder_contents(request, folder_id):
    try:
        # Get the client inventory folder by ID
        folder = ClientInventoryFolder.objects.get(id=folder_id)
        transactions = InventoryHistory.objects.filter(client_inventory_folder=folder)

        # Create a workbook and a sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Client Inventory Records'

        # Define header styles
        header_font = Font(bold=True)
        gold_fill = PatternFill(start_color='F59B00', end_color='00B0F0', fill_type='solid')
        currency_format = '#,##0.00'

        # Define the headers
        headers = [
            'Transaction ID', 'Client', 'Date', 'Item', 'Quantity In', 'Quantity Out',
            'Price', 'Total Amount', 'Delivery Ref', 'Delivery No.', 'Invoice Type', 'Invoice#'
        ]
        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = gold_fill
            cell.value = cell.value.upper() if cell.value is not None else cell.value

        # Populate the sheet with data
        for transaction in transactions:
            sheet.append([
                transaction.id,
                transaction.client if transaction.client else 'N/A',  # Assuming client is a string or related field
                transaction.date.strftime('%Y-%m-%d') if transaction.date else 'N/A',
                str(transaction.item),  # Use 'item' if that's the correct field
                transaction.quantity_in,
                transaction.quantity_out,
                transaction.price,
                transaction.total_amount,
                transaction.delivery_ref,
                transaction.delivery_no,
                transaction.invoice_type,
                transaction.invoice_no,
            ])

        # Apply currency format to price and total_amount columns
        for cell in sheet['G']:  # Assuming price is in column G (index 7)
            if cell.row > 1:  # Skip header row
                cell.number_format = currency_format

        for cell in sheet['H']:  # Assuming total_amount is in column H (index 8)
            if cell.row > 1:  # Skip header row
                cell.number_format = currency_format

        # Adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Add extra space for better visibility
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Create an in-memory buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Set the response to return the Excel file
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=ClientInventory_{folder.name}.xlsx'
        return response

    except ClientInventoryFolder.DoesNotExist:
        return HttpResponse("Client Inventory Folder not found", status=404)


def export_all_client_folders(request):
    # Create an in-memory buffer to hold the zip file
    buffer = BytesIO()

    # Create a zip file in the buffer
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Iterate through each client folder
        for folder in ClientInventoryFolder.objects.all():
            # Get inventory records for the folder
            transactions = InventoryHistory.objects.filter(client_inventory_folder=folder)

            # Create a workbook and a sheet
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'Client Inventory Records'

            # Define header styles
            header_font = Font(bold=True)
            gold_fill = PatternFill(start_color='F59B00', end_color='F59B00', fill_type='solid')
            currency_format = '#,##0.00'

            # Define the headers
            headers = [
                'Transaction ID', 'Client', 'Date', 'Item', 'Quantity In', 'Quantity Out',
                'Price', 'Total Amount', 'Delivery Ref', 'Delivery No.', 'Invoice Type', 'Invoice#'
            ]
            sheet.append(headers)

            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = gold_fill
                cell.value = cell.value.upper() if cell.value is not None else cell.value

            # Populate the sheet with data
            for transaction in transactions:
                sheet.append([
                    transaction.id,
                    transaction.client if transaction.client else 'N/A',  # Assuming client is a string or related field
                    transaction.date.strftime('%Y-%m-%d') if transaction.date else 'N/A',
                    str(transaction.item),  # Use 'item' if that's the correct field
                    transaction.quantity_in,
                    transaction.quantity_out,
                    transaction.price,
                    transaction.total_amount,
                    transaction.delivery_ref,
                    transaction.delivery_no,
                    transaction.invoice_type,
                    transaction.invoice_no,
                ])

            # Apply currency format to price and total_amount columns
            for cell in sheet['G']:  # Assuming price is in column G (index 7)
                if cell.row > 1:  # Skip header row
                    cell.number_format = currency_format

            for cell in sheet['H']:  # Assuming total_amount is in column H (index 8)
                if cell.row > 1:  # Skip header row
                    cell.number_format = currency_format

            # Adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)  # Add extra space for better visibility
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook to a bytes buffer
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            # Add the workbook to the zip file
            zf.writestr(f'ClientInventory_{folder.name}.xlsx', excel_buffer.getvalue())

    buffer.seek(0)

    # Set the response to return the zip file
    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=client_inventory_folders.zip'
    return response


# ----------------------------Transaction History----------------------------------------------
def transaction_history(request):
    query = request.GET.get('q')  # Get search query from request
    page_number = request.GET.get('page', 1)  # Get page number from request

    # Dictionary to map full and abbreviated month names to month numbers
    month_mapping = {
        "January": "1", "February": "2", "March": "3", "April": "4",
        "May": "5", "June": "6", "July": "7", "August": "8",
        "September": "9", "October": "10", "November": "11", "December": "12",
        "Jan": "1", "Feb": "2", "Mar": "3", "Apr": "4",
        "Jun": "6", "Jul": "7", "Aug": "8", "Sep": "9", "Oct": "10", "Nov": "11", "Dec": "12"
    }

    # Retrieve all records from InventoryHistory and exclude records where date is null
    transactions = InventoryHistory.objects.exclude(date__isnull=True).order_by('-date')

    # Apply search filter if a query is present
    if query:
        try:
            # Try to interpret the query as a full date (e.g., 'Aug 20, 2024')
            date_obj = datetime.strptime(query, '%b %d, %Y').date()
            transactions = transactions.filter(date=date_obj)
        except ValueError:
            # If it's not a full date, check if the query is a month name (full or abbreviated)
            for month_name, month_number in month_mapping.items():
                if month_name.lower() in query.lower():
                    # Check if the query contains a year
                    year = None
                    try:
                        year = int(query.split()[-1])  # Try to extract the year
                    except (ValueError, IndexError):
                        pass

                    # Filter based on month and possibly year
                    if year:
                        transactions = transactions.filter(date__month=month_number, date__year=year)
                    else:
                        transactions = transactions.filter(date__month=month_number)
                    break
            else:
                # If not a date or month, treat as a general text search
                transactions = transactions.filter(
                    Q(item_code__icontains=query) |
                    Q(supplier__icontains=query) |
                    Q(po_product_name__icontains=query) |
                    Q(new_product_name__icontains=query) |
                    Q(unit__icontains=query) |
                    Q(quantity_out__icontains=query) |
                    Q(price__icontains=query) |
                    Q(total_amount__icontains=query) |
                    Q(site_delivered__icontains=query) |
                    Q(client__icontains=query)
                )

    # Paginate the filtered transactions
    paginator = Paginator(transactions, 100)  # Show 100 transactions per page
    try:
        transactions_page = paginator.page(page_number)
    except PageNotAnInteger:
        transactions_page = paginator.page(1)
    except EmptyPage:
        transactions_page = paginator.page(paginator.num_pages)

    return render(request, 'Inventory/transaction_history.html', {
        'transactions': transactions_page,
        'query': query,
        'page_number': page_number
    })


def create_site_inventory_folder(request):
    if request.method == 'POST':
        folder_name = request.POST.get('folder_name')

        if folder_name:
            new_folder, created = SiteInventoryFolder.objects.get_or_create(name=folder_name)

            if created:
                # Handle matching transactions here
                matching_transactions = InventoryHistory.objects.filter(site_delivered=folder_name)
                for transaction in matching_transactions:
                    transaction.site_inventory_folder = new_folder
                    transaction.save()

                return JsonResponse({'success': True, 'message': 'Folder created successfully.'})
            else:
                return JsonResponse({'success': False, 'message': 'Folder with this name already exists.'})

    return JsonResponse({'success': False, 'error': 'Invalid request method.'})


def delete_site_inventory_folder(request, folder_id):
    if request.method == 'POST':
        folder = get_object_or_404(SiteInventoryFolder, id=folder_id)
        folder.delete()  # This will set the site_inventory_folder field in InventoryHistory to NULL
        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Invalid request method.'})


def site_inventory_folder_list(request):
    if request.method == 'POST':
        folder_name = request.POST.get('folder_name')

        if folder_name:
            # Create or get the folder
            new_folder, created = SiteInventoryFolder.objects.get_or_create(name=folder_name)

            if created:
                # Removed code:
                # The following code was responsible for finding and associating matching PurchaseOrder records
                # with the newly created SiteInventoryFolder:
                #
                # matching_orders = PurchaseOrder.objects.filter(site_delivered=folder_name)
                # for order in matching_orders:
                #     order.site_inventory_folder = new_folder
                #     order.save()

                return JsonResponse({'success': True, 'message': 'Folder created successfully.'})
            else:
                return JsonResponse({'success': False, 'message': 'Folder with this name already exists.'})

    # Handling GET requests to render the folder list
    folders = SiteInventoryFolder.objects.all()
    context = {
        'folders': folders,
    }
    return render(request, 'Inventory/site_inventory_folder_list.html', context)


def view_site_inventory_folder_contents(request, folder_id):
    folder = get_object_or_404(SiteInventoryFolder, id=folder_id)
    transactions_list = InventoryHistory.objects.filter(site_inventory_folder=folder)

    # Calculate total_amount
    total_amount = transactions_list.aggregate(total_amount_sum=Sum('total_amount'))['total_amount_sum'] or 0

    # Pagination
    paginator = Paginator(transactions_list, 100)  # Show 20 transactions per page
    page_number = request.GET.get('page')
    try:
        transactions = paginator.page(page_number)
    except PageNotAnInteger:
        transactions = paginator.page(1)
    except EmptyPage:
        transactions = paginator.page(paginator.num_pages)

    context = {
        'folder': folder,
        'transactions': transactions,
        'total_amount': total_amount,
    }

    return render(request, 'Inventory/site_folder_contents.html', context)


def create_client_inventory_folder(request):
    if request.method == 'POST':
        folder_name = request.POST.get('folder_name')

        if folder_name:
            new_folder, created = ClientInventoryFolder.objects.get_or_create(name=folder_name)

            if created:
                # Handle matching transactions here
                matching_transactions = InventoryHistory.objects.filter(client=folder_name)
                for transaction in matching_transactions:
                    transaction.client_inventory_folder = new_folder
                    transaction.save()

                return JsonResponse(
                    {'success': True, 'message': 'Client folder created and transactions updated successfully.'})
            else:
                return JsonResponse({'success': False, 'message': 'Client folder with this name already exists.'})

    return JsonResponse({'success': False, 'error': 'Invalid request method.'})


def delete_client_inventory_folder(request, folder_id):
    if request.method == 'POST':
        folder = get_object_or_404(ClientInventoryFolder, id=folder_id)
        folder.delete()  # This will set the client_inventory_folder field in InventoryHistory to NULL
        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Invalid request method.'})


def client_inventory_folder_list(request):
    if request.method == 'POST':
        folder_name = request.POST.get('folder_name')

        if folder_name:
            # Create or get the folder
            new_folder, created = ClientInventoryFolder.objects.get_or_create(name=folder_name)

            if created:
                # Removed code:
                # The following code was responsible for finding and associating matching PurchaseOrder records
                # with the newly created ClientInventoryFolder:
                #
                # matching_orders = PurchaseOrder.objects.filter(client=folder_name)
                # for order in matching_orders:
                #     order.client_inventory_folder = new_folder
                #     order.save()

                return JsonResponse({'success': True, 'message': 'Folder created successfully.'})
            else:
                return JsonResponse({'success': False, 'message': 'Folder with this name already exists.'})

    # Handling GET requests to render the folder list
    folders = ClientInventoryFolder.objects.all()
    context = {
        'folders': folders,
    }
    return render(request, 'Inventory/client_inventory_folder_list.html', context)


def view_client_inventory_folder_contents(request, folder_id):
    folder = get_object_or_404(ClientInventoryFolder, id=folder_id)
    transactions_list = InventoryHistory.objects.filter(client_inventory_folder=folder)

    total_amount = transactions_list.aggregate(total_amount_sum=Sum('total_amount'))['total_amount_sum'] or 0

    # Pagination
    paginator = Paginator(transactions_list, 100)  # Show 20 transactions per page
    page_number = request.GET.get('page')
    try:
        transactions = paginator.page(page_number)
    except PageNotAnInteger:
        transactions = paginator.page(1)
    except EmptyPage:
        transactions = paginator.page(paginator.num_pages)

    context = {
        'folder': folder,
        'transactions': transactions,
        'total_amount': total_amount,
    }

    return render(request, 'Inventory/client_folder_contents.html', context)


def edit_inventory_history_remarks(request, folder_id, record_id):
    # Get the folder and record
    folder = get_object_or_404(ClientInventoryFolder, id=folder_id)
    record = get_object_or_404(InventoryHistory, client_inventory_folder=folder, id=record_id)

    if request.method == 'POST':
        form = EditRemarksForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            return redirect('view_client_inventory_folder_contents', folder_id=folder_id)  # Redirect to folder view after edit
    else:
        form = EditRemarksForm(instance=record)

    context = {
        'form': form,
        'folder': folder,
        'record': record
    }
    return render(request, 'Inventory/client_folder_contents.html', context)

@csrf_exempt
def edit_remarks(request, transaction_id):
    if request.method == 'POST':
        data = json.loads(request.body)
        remarks = data.get('remarks', '')

        try:
            transaction = InventoryHistory.objects.get(id=transaction_id)
            transaction.remarks = remarks
            transaction.save()
            return JsonResponse({'success': True})
        except InventoryHistory.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Transaction not found'}, status=404)

    return JsonResponse({'success': False}, status=400)


@csrf_exempt
def edit_site_remarks(request, transaction_id):
    if request.method == 'POST':
        data = json.loads(request.body)
        remarks = data.get('remarks', '')

        try:
            transaction = InventoryHistory.objects.get(id=transaction_id)
            transaction.remarks = remarks
            transaction.save()
            return JsonResponse({'success': True})
        except InventoryHistory.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Transaction not found'}, status=404)

    return JsonResponse({'success': False}, status=400)


# ----------------------------SUPPLIER----------------------------------------------

def delete_supplier_folder(request, folder_id):
    if request.method == 'POST':
        folder = get_object_or_404(SupplierFolder, id=folder_id)
        folder.delete()  # This will set the supplier_folder field in PurchaseOrder to NULL
        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Invalid request method.'})


def supplier_list_folders(request):
    if request.method == 'POST':
        folder_name = request.POST.get('folder_name')

        if folder_name:
            new_folder, created = SupplierFolder.objects.get_or_create(name=folder_name)

            if created:
                # Handle matching orders here
                matching_orders = PurchaseOrder.objects.filter(supplier=folder_name)
                for order in matching_orders:
                    order.supplier_folder = new_folder
                    order.save()

                return JsonResponse({'success': True, 'message': 'Folder created successfully.'})
            else:
                return JsonResponse({'success': False, 'message': 'Folder with this name already exists.'})

    # Handling GET requests to render the folder list
    folders = SupplierFolder.objects.all()
    context = {
        'folders': folders,
    }
    return render(request, 'supplier/supplier_list_folders.html', context)


def view_folder_contents(request, folder_id):
    folder = get_object_or_404(SupplierFolder, id=folder_id)

    # Get the search query from the GET parameters
    query = request.GET.get('q', '').strip()
    page_number = request.GET.get('page', 1)

    # Initial queryset
    purchase_orders = PurchaseOrder.objects.filter(supplier_folder=folder)

    # If there's a search query, filter the purchase orders accordingly
    if query:
        try:
            # Try to interpret the query as a month number (e.g., "08" for August)
            month_number = int(query)
            purchase_orders = purchase_orders.filter(date__month=month_number)
        except ValueError:
            try:
                # Try to interpret the query as a full month name only (e.g., "August")
                month_mapping = {
                    "January": "1", "February": "2", "March": "3", "April": "4",
                    "May": "5", "June": "6", "July": "7", "August": "8",
                    "September": "9", "October": "10", "November": "11", "December": "12"
                }
                month_number = month_mapping.get(query.capitalize())
                if month_number:
                    purchase_orders = purchase_orders.filter(date__month=month_number)
                else:
                    raise ValueError("Not a valid month name")
            except ValueError:
                # If not a date, treat it as a general text search
                purchase_orders = purchase_orders.filter(
                    Q(po_number__icontains=query) |
                    Q(purchaser__icontains=query) |
                    Q(brand__icontains=query) |
                    Q(item_code__icontains=query) |
                    Q(particulars__icontains=query) |
                    Q(quantity__icontains=query) |
                    Q(unit__icontains=query) |
                    Q(price__icontains=query) |
                    Q(total_amount__icontains=query) |
                    Q(site_delivered__icontains=query) |
                    Q(fbbd_ref_number__icontains=query) |
                    Q(remarks__icontains=query) |
                    Q(supplier__icontains=query) |
                    Q(delivery_ref__icontains=query) |
                    Q(delivery_no__icontains=query) |
                    Q(invoice_type__icontains=query) |
                    Q(invoice_no__icontains=query) |
                    Q(payment_req_ref__icontains=query) |
                    Q(payment_details__icontains=query) |
                    Q(remarks2__icontains=query)
                )

    # Pagination
    paginator = Paginator(purchase_orders, 100)  # Show 20 orders per page
    try:
        orders = paginator.page(page_number)
    except PageNotAnInteger:
        orders = paginator.page(1)
    except EmptyPage:
        orders = paginator.page(paginator.num_pages)

    # Render the template with the context
    context = {
        'folder': folder,
        'purchase_orders': orders,
        'query': query,
        'page_number': page_number,
    }
    return render(request, 'supplier/view_folder_contents.html', context)


# -----------------------------SHOP----------------------------------------------

def bulk_edit_inventory(request):
    if request.method == 'POST':
        if 'add_to_cart' in request.POST:
            item_ids = request.POST.getlist('item_ids')

            # Ensure there are selected items
            success = False
            for item_id in item_ids:
                quantity_out = request.POST.get(f'quantity_out_{item_id}', '0')
                quantity_out = Decimal(quantity_out) if quantity_out.strip() != '' else 0

                if quantity_out > 0:
                    success = True  # At least one valid item has been edited
                    break

            if not success:
                messages.error(request,
                               'No valid quantity out entered for any item. Please enter a quantity to add to the cart.')
                return redirect('bulk_edit_inventory')

            success = True
            errors = []

            for item_id in item_ids:
                quantity_out = request.POST.get(f'quantity_out_{item_id}', '0')
                quantity_out = Decimal(quantity_out) if quantity_out.strip() != '' else 0

                if quantity_out > 0:
                    try:
                        item = ItemInventory.objects.get(id=item_id)
                        cart_item, created = Cart.objects.get_or_create(
                            item=item,
                            defaults={'quantity': quantity_out}
                        )
                        if not created:
                            cart_item.quantity += quantity_out
                            cart_item.save()

                    except ItemInventory.DoesNotExist:
                        success = False
                        errors.append(f"Item with ID {item_id} does not exist.")
                    except Exception as e:
                        success = False
                        errors.append(f"An error occurred for item ID {item_id}: {str(e)}")

            if success:
                messages.success(request, 'Items added to cart successfully.')
            else:
                messages.error(request, 'Some errors occurred: ' + ', '.join(errors))
            return redirect('bulk_edit_inventory')

        elif 'finalize_changes' in request.POST:
            # Check if the cart is empty before processing
            cart_items = Cart.objects.all()
            if not cart_items.exists():
                messages.error(request, 'No items in the cart to finalize.')
                return redirect('bulk_edit_inventory')

            form = ItemInventoryBulkForm(request.POST)
            if form.is_valid():
                date = form.cleaned_data['date']
                location_type = form.cleaned_data['location_type']
                location_name = form.cleaned_data['location_name']
                delivery_ref = form.cleaned_data['delivery_ref']
                delivery_no = form.cleaned_data['delivery_no']
                invoice_type = form.cleaned_data['invoice_type']
                invoice_no = form.cleaned_data['invoice_no']

                success = True
                errors = []

                for cart_item in cart_items:
                    try:
                        item = cart_item.item
                        quantity_out = cart_item.quantity
                        price = item.price
                        total_amount = quantity_out * price

                        # Determine folder based on location_type
                        if location_type == 'site':
                            folder, created = SiteInventoryFolder.objects.get_or_create(name=location_name)
                            item.site_inventory_folder = folder
                            item.site_delivered = location_name

                        elif location_type == 'client':
                            client_folder, created = ClientInventoryFolder.objects.get_or_create(name=location_name)
                            item.client_inventory_folder = client_folder
                            item.client = location_name

                        # Update ItemInventory details
                        item.date = date
                        item.quantity_out += quantity_out
                        item.stock = item.quantity_in - item.quantity_out
                        item.total_amount = total_amount
                        item.delivery_ref = delivery_ref
                        item.delivery_no = delivery_no
                        item.invoice_type = invoice_type
                        item.invoice_no = invoice_no
                        item.site_or_client_choice = location_type

                        item.save()

                        # Create an InventoryHistory record
                        InventoryHistory.objects.create(
                            item=item,
                            date=date,
                            item_code=item.item_code,
                            supplier=item.supplier,
                            po_product_name=item.po_product_name,
                            unit=item.unit,
                            quantity_in=item.quantity_in,
                            quantity_out=quantity_out,
                            stock=item.stock,
                            price=price,
                            total_amount=total_amount,
                            delivery_ref=delivery_ref,
                            delivery_no=delivery_no,
                            invoice_type=invoice_type,
                            invoice_no=invoice_no,
                            site_inventory_folder=item.site_inventory_folder if location_type == 'site' else None,
                            client_inventory_folder=item.client_inventory_folder if location_type == 'client' else None,
                            site_delivered=item.site_delivered if location_type == 'site' else None,
                            client=item.client if location_type == 'client' else None,
                        )

                        # Remove the item from the cart
                        cart_item.delete()

                    except ItemInventory.DoesNotExist:
                        success = False
                        errors.append(f"Item with ID {cart_item.item.id} does not exist.")
                    except Exception as e:
                        success = False
                        errors.append(f"An error occurred for item ID {cart_item.item.id}: {str(e)}")

                # Clear the cart after processing
                Cart.objects.all().delete()

                if success:
                    messages.success(request, 'Items updated successfully.')
                else:
                    messages.error(request, 'Some errors occurred: ' + ', '.join(errors))
                return redirect('bulk_edit_inventory')

    else:
        query = request.GET.get('q', '')
        if query:
            items = ItemInventory.objects.filter(po_product_name__icontains=query)
        else:
            items = ItemInventory.objects.all()

        cart_items = Cart.objects.all()
        for cart_item in cart_items:
            cart_item.total_amount = cart_item.quantity * cart_item.item.price

        # Ensure the form is initialized regardless of the query
        form = ItemInventoryBulkForm()

    return render(request, 'Inventory/bulk_edit_inventory.html', {
        'form': form,
        'items': items,
        'cart_items': cart_items,
        'query': query,
    })



def remove_cart_item(request, cart_item_id):
    cart_item = get_object_or_404(Cart, id=cart_item_id)
    cart_item.delete()
    messages.success(request, 'Item removed from cart successfully.')
    return redirect('bulk_edit_inventory')


def bulk_edit_purchase_order(request):
    if request.method == 'POST':
        if 'add_to_cart' in request.POST:
            po_ids = request.POST.getlist('po_ids')
            success = True
            errors = []

            # Add selected items to the cart
            for po_id in po_ids:
                try:
                    po = PurchaseOrder.objects.get(id=po_id)
                    cart_item, created = poCart.objects.get_or_create(
                        particulars=po,
                        defaults={'fbbd_ref_number': '', 'remarks2': ''}
                    )
                except PurchaseOrder.DoesNotExist:
                    success = False
                    errors.append(f"Purchase Order with ID {po_id} does not exist.")
                except Exception as e:
                    success = False
                    errors.append(f"An error occurred for Purchase Order ID {po_id}: {str(e)}")

            if success:
                messages.success(request, 'Orders added to cart successfully.')
            else:
                messages.error(request, 'Some errors occurred: ' + ', '.join(errors))
            return redirect('bulk_edit_purchase_order')

        elif 'finalize_changes' in request.POST:
            if poCart.objects.count() == 0:
                messages.error(request, 'The cart is empty. Please add items to the cart before finalizing changes.')
                return redirect('bulk_edit_purchase_order')

            fbbd_ref_number = request.POST.get('fbbd_ref_number', '')
            remarks2 = request.POST.get('remarks2', '')
            success = True
            errors = []
            changes_made = False

            cart_items = poCart.objects.all()

            for cart_item in cart_items:
                po = cart_item.particulars
                # Check if the values have actually changed
                if (po.fbbd_ref_number != fbbd_ref_number or po.remarks2 != remarks2):
                    changes_made = True
                    po.fbbd_ref_number = fbbd_ref_number
                    po.remarks2 = remarks2
                    try:
                        po.save()
                    except Exception as e:
                        success = False
                        errors.append(f"An error occurred while saving Purchase Order ID {po.id}: {str(e)}")
                    cart_item.delete()

            if not changes_made:
                messages.error(request,
                               'No changes were made to any items. Please edit at least one item before finalizing changes.')
                return redirect('bulk_edit_purchase_order')

            # Clear the cart after finalizing changes
            if success:
                poCart.objects.all().delete()
                messages.success(request, 'Orders updated successfully.')
            else:
                messages.error(request, 'Some errors occurred: ' + ', '.join(errors))

            return redirect('bulk_edit_purchase_order')

    else:
        query = request.GET.get('q', '')
        if query:
            orders = PurchaseOrder.objects.filter(
                Q(po_number__icontains=query) |
                Q(particulars__icontains=query) |
                Q(quantity__icontains=query) |
                Q(price__icontains=query)
            )
        else:
            orders = PurchaseOrder.objects.all()

        form = PurchaseOrderBulkForm()

        return render(request, 'records/purchase_order_update.html', {
            'form': form,
            'orders': orders,
            'cart_items': poCart.objects.all(),
            'query': query,
            'remarks2_choices': PurchaseOrderBulkForm.REMARKS2_CHOICES,
        })


def po_remove_cart_item(request, item_id):
    try:
        cart_item = get_object_or_404(poCart, id=item_id)
        cart_item.delete()
        messages.success(request, 'Item removed from cart successfully.')
    except poCart.DoesNotExist:
        messages.error(request, 'Item does not exist.')
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')

    return redirect('bulk_edit_purchase_order')


def remove_all_cart_items(request):
    try:
        poCart.objects.all().delete()
        messages.success(request, 'All items removed from cart successfully.')
    except Exception as e:
        messages.error(request, f'An error occurred while removing items from the cart: {str(e)}')

    return redirect('bulk_edit_purchase_order')


# -----------------------------STOCK IN INVENTORY----------------------------------------------


def stock_in_create(request):
    if request.method == 'POST':
        form = StockInHistoryForm(request.POST)
        if form.is_valid():
            # Save the new stock in history but don't commit to DB yet
            stock_in = form.save(commit=False)

            # Check if the particulars match an existing po_product_name in ItemInventory
            if not ItemInventory.objects.filter(po_product_name=stock_in.particulars).exists():
                # Raise an error if the item does not exist in the ItemInventory list
                return JsonResponse({'status': 'error',
                                     'message': 'The item does not exist in the item list. Please check the particulars.'})

            # Get the supplier name from the stock in history
            supplier_name = stock_in.supplier

            if supplier_name:
                # Create or get the InventorySupplierFolder associated with this supplier
                folder, created = InventorySupplierFolder.objects.get_or_create(name=supplier_name)

                # Associate the new stock in history with the supplier's folder
                stock_in.supplier_folder = folder
                stock_in.save()

                # Ensure all existing stock in histories with the same supplier are associated with the folder
                matching_histories = StockInHistory.objects.filter(supplier=supplier_name, supplier_folder__isnull=True)
                for history in matching_histories:
                    history.supplier_folder = folder
                    history.save()

            # Update or create the related ItemInventory record
            item_inventory, created = ItemInventory.objects.get_or_create(
                item_code=stock_in.item_code,
                supplier=stock_in.supplier,
                defaults={
                    'po_product_name': stock_in.particulars,
                    'unit': stock_in.unit,
                    'quantity_in': stock_in.quantity_in,
                    'quantity_out': 0,  # Set quantity_out to 0 if the record is new
                    'stock': stock_in.quantity_in  # Initialize stock to quantity_in for new items
                }
            )

            if not created:
                # Update the existing ItemInventory record by adding the quantity_in
                item_inventory.quantity_in += stock_in.quantity_in
                item_inventory.stock = item_inventory.quantity_in - item_inventory.quantity_out
                item_inventory.quantity_out = item_inventory.quantity_out or 0  # Ensure quantity_out is never None
                item_inventory.save()

            return JsonResponse({'status': 'success'})
        else:
            print(form.errors)  # Debugging form errors
            return JsonResponse({'status': 'error', 'errors': form.errors})
    else:
        form = StockInHistoryForm()

    return render(request, 'Inventory/stockIn/stock_in_form.html', {'form': form})


def delete_inventory_supplier_folder(request, folder_id):
    if request.method == 'POST':
        folder = get_object_or_404(InventorySupplierFolder, id=folder_id)
        folder.delete()  # This will set the supplier_folder field in PurchaseOrder to NULL
        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Invalid request method.'})


def inventory_supplier_list_folders(request):
    if request.method == 'POST':
        folder_name = request.POST.get('folder_name')

        if folder_name:
            # Create or get the InventorySupplierFolder based on the folder_name
            new_folder, created = InventorySupplierFolder.objects.get_or_create(name=folder_name)

            if created:
                # Handle matching stock in records here
                matching_stock_ins = StockInHistory.objects.filter(supplier=folder_name)
                for stock_in in matching_stock_ins:
                    stock_in.supplier_folder = new_folder
                    stock_in.save()

                return JsonResponse({'success': True, 'message': 'Folder created successfully.'})
            else:
                return JsonResponse({'success': False, 'message': 'Folder with this name already exists.'})

    # Handling GET requests to render the folder list
    folders = InventorySupplierFolder.objects.all()
    context = {
        'folders': folders,
    }
    return render(request, 'Inventory/stockIn/inventory_supplier_list_folders.html', context)


def inventory_supplier_contents(request, folder_id):
    folder = get_object_or_404(InventorySupplierFolder, id=folder_id)

    # Get the search query from the GET parameters
    query = request.GET.get('q', '').strip()
    page_number = request.GET.get('page', 1)

    # Initial queryset
    stock_in_histories = StockInHistory.objects.filter(supplier_folder=folder)

    # If there's a search query, filter the stock in records accordingly
    if query:
        try:
            # Try to interpret the query as a month number (e.g., "08" for August)
            month_number = int(query)
            stock_in_histories = stock_in_histories.filter(date__month=month_number)
        except ValueError:
            try:
                # Try to interpret the query as a full month name only (e.g., "August")
                month_mapping = {
                    "January": "1", "February": "2", "March": "3", "April": "4",
                    "May": "5", "June": "6", "July": "7", "August": "8",
                    "September": "9", "October": "10", "November": "11", "December": "12"
                }
                month_number = month_mapping.get(query.capitalize())
                if month_number:
                    stock_in_histories = stock_in_histories.filter(date__month=month_number)
                else:
                    raise ValueError("Not a valid month name")
            except ValueError:
                # If not a date, treat it as a general text search
                stock_in_histories = stock_in_histories.filter(
                    Q(po_number__icontains=query) |
                    Q(purchaser__icontains=query) |
                    Q(item_code__icontains=query) |
                    Q(particulars__icontains=query) |
                    Q(quantity_in__icontains=query) |
                    Q(unit__icontains=query) |
                    Q(fbbd_ref_number__icontains=query) |
                    Q(remarks__icontains=query) |
                    Q(supplier__icontains=query) |
                    Q(delivery_ref__icontains=query) |
                    Q(delivery_no__icontains=query) |
                    Q(invoice_type__icontains=query) |
                    Q(invoice_no__icontains=query) |
                    Q(payment_req_ref__icontains=query) |
                    Q(payment_details__icontains=query) |
                    Q(remarks2__icontains=query)
                )

    # Pagination
    paginator = Paginator(stock_in_histories, 100)  # Show 100 records per page
    try:
        stock_in_records = paginator.page(page_number)
    except PageNotAnInteger:
        stock_in_records = paginator.page(1)
    except EmptyPage:
        stock_in_records = paginator.page(paginator.num_pages)

    # Render the template with the context
    context = {
        'folder': folder,
        'stock_in_records': stock_in_records,
        'query': query,
        'page_number': page_number,
    }
    return render(request, 'Inventory/stockIn/inventory_supplier_contents.html', context)


def export_inventory_supplier_contents(request, folder_id):
    try:
        # Get the InventorySupplierFolder by ID
        folder = InventorySupplierFolder.objects.get(id=folder_id)
        stock_in_list = StockInHistory.objects.filter(supplier_folder=folder)

        # Create a workbook and a sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Stock In Records'

        # Define header styles
        header_font = Font(bold=True)
        blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        currency_format = '#,##0.00'

        # Define the headers
        headers = [
            'Date', 'PO Number', 'Purchaser', 'Item Code', 'Particular',
            'Quantity In', 'Unit', 'FBBD Ref#', 'Remarks', 'Supplier',
            'Delivery Ref#', 'Delivery No.', 'Invoice Type', 'Invoice No.',
            'Payment Req Ref#', 'Payment Details', 'Remarks2'
        ]
        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = blue_fill
            cell.value = cell.value.upper() if cell.value is not None else cell.value

        # Populate the sheet with data
        for stock_in in stock_in_list:
            sheet.append([
                stock_in.date.strftime('%Y-%m-%d') if stock_in.date else 'N/A',
                stock_in.po_number,
                stock_in.purchaser,
                stock_in.item_code,
                stock_in.particulars,
                stock_in.quantity_in,
                stock_in.unit,
                stock_in.fbbd_ref_number,
                stock_in.remarks,
                stock_in.supplier,
                stock_in.delivery_ref,
                stock_in.delivery_no,
                stock_in.invoice_type,
                stock_in.invoice_no,
                stock_in.payment_req_ref,
                stock_in.payment_details,
                stock_in.remarks2
            ])

        for cell in sheet['F']:  # Assuming quantity_in is in column F (index 6)
            if cell.row > 1:  # Skip header row
                cell.number_format = currency_format

        # Auto-adjust column widths based on content
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Add extra space for better visibility
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Create an in-memory buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Set the response to return the Excel file
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=StockInRecords_{folder.name}.xlsx'
        return response

    except InventorySupplierFolder.DoesNotExist:
        return HttpResponse("Inventory Supplier Folder not found", status=404)


def export_stock_in_transaction_history_to_excel(request):
    query = request.GET.get('q')  # Search query parameter
    date_query = request.GET.get('date')  # Date query parameter

    # Start with all stock-in transactions
    transactions = StockInHistory.objects.exclude(date__isnull=True).order_by('-date')

    # Apply search filter
    if query:
        transactions = transactions.filter(
            Q(date__icontains=query) |
            Q(item_code__icontains=query) |
            Q(supplier__icontains=query) |
            Q(particulars__icontains=query) |
            Q(unit__icontains=query) |
            Q(quantity_in__icontains=query) |
            Q(invoice_no__icontains=query) |
            Q(invoice_type__icontains=query)
        )

    # Apply date filter
    if date_query:
        try:
            date_obj = datetime.strptime(date_query, '%b %d, %Y').date()
            transactions = transactions.filter(date=date_obj)
        except ValueError:
            try:
                date_obj = datetime.strptime(date_query, '%b %Y')
                transactions = transactions.filter(date__year=date_obj.year, date__month=date_obj.month)
            except ValueError:
                try:
                    date_obj = datetime.strptime(date_query, '%B %Y')
                    transactions = transactions.filter(date__year=date_obj.year, date__month=date_obj.month)
                except ValueError:
                    try:
                        month_number = int(date_query)
                        transactions = transactions.filter(date__month=month_number)
                    except ValueError:
                        try:
                            date_obj = datetime.strptime(date_query, '%B')
                            transactions = transactions.filter(date__month=date_obj.month)
                        except ValueError:
                            pass

    # Create a workbook and a sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Stock In Transaction History'

    header_font = Font(color='FFFFFF', bold=True)
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    # Define the headers
    headers = [
        'Date', 'PO#','Purchaser', 'Item Code', 'Particular', 'Unit',
        'Quantity In', 'Supplier', 'Remarks', 'Invoice No.', 'Invoice Type'
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = black_fill
        cell.value = cell.value.upper() if cell.value is not None else cell.value

    # Populate the sheet with data
    for transaction in transactions:
        sheet.append([
            transaction.date.strftime('%Y-%m-%d') if transaction.date else 'N/A',
            transaction.po_number, transaction.purchaser, transaction.item_code, transaction.particulars,
            transaction.unit, transaction.quantity_in, transaction.supplier,
            transaction.remarks, transaction.invoice_no, transaction.invoice_type
        ])

    # Auto-size columns for better visibility
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Create an in-memory buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Set the response to return the Excel file
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=StockInTransactionHistory.xlsx'
    return response


def get_item_details(request):
    particulars = request.GET.get('particulars', None)

    if particulars:
        try:
            item = ItemInventory.objects.get(po_product_name=particulars)
            return JsonResponse({
                'item_code': item.item_code,
                'unit': item.unit,
                'supplier': item.supplier,
            })
        except ItemInventory.DoesNotExist:
            return JsonResponse({'error': 'Item not found'}, status=404)

    return JsonResponse({'error': 'Invalid request'}, status=400)


def get_item_inventory(request):
    items = list(ItemInventory.objects.values('id', 'po_product_name'))
    return JsonResponse({
        'items': [{'id': item['id'], 'text': item['po_product_name']} for item in items]
    })


def stock_in_transaction_history(request):
    query = request.GET.get('q')  # Get search query from request
    page_number = request.GET.get('page', 1)  # Get page number from request

    # Dictionary to map full and abbreviated month names to month numbers
    month_mapping = {
        "January": "1", "February": "2", "March": "3", "April": "4",
        "May": "5", "June": "6", "July": "7", "August": "8",
        "September": "9", "October": "10", "November": "11", "December": "12",
        "Jan": "1", "Feb": "2", "Mar": "3", "Apr": "4",
        "Jun": "6", "Jul": "7", "Aug": "8", "Sep": "9", "Oct": "10", "Nov": "11", "Dec": "12"
    }

    # Retrieve all records from StockInHistory and exclude records where date is null
    stock_in_transactions = StockInHistory.objects.exclude(date__isnull=True).order_by('-date')

    # Apply search filter if a query is present
    if query:
        try:
            # Try to interpret the query as a full date (e.g., 'Aug 20, 2024')
            date_obj = datetime.strptime(query, '%b %d, %Y').date()
            stock_in_transactions = stock_in_transactions.filter(date=date_obj)
        except ValueError:
            # If it's not a full date, check if the query is a month name (full or abbreviated)
            for month_name, month_number in month_mapping.items():
                if month_name.lower() in query.lower():
                    # Check if the query contains a year
                    year = None
                    try:
                        year = int(query.split()[-1])  # Try to extract the year
                    except (ValueError, IndexError):
                        pass

                    # Filter based on month and possibly year
                    if year:
                        stock_in_transactions = stock_in_transactions.filter(date__month=month_number, date__year=year)
                    else:
                        stock_in_transactions = stock_in_transactions.filter(date__month=month_number)
                    break
            else:
                # If not a date or month, treat as a general text search
                stock_in_transactions = stock_in_transactions.filter(
                    Q(item_code__icontains=query) |
                    Q(supplier__icontains=query) |
                    Q(particulars__icontains=query) |
                    Q(unit__icontains=query) |
                    Q(quantity_in__icontains=query) |
                    Q(remarks__icontains=query) |
                    Q(invoice_no__icontains=query) |
                    Q(po_number__icontains=query)
                )

    # Paginate the filtered stock-in transactions
    paginator = Paginator(stock_in_transactions, 100)  # Show 100 transactions per page
    try:
        stock_in_page = paginator.page(page_number)
    except PageNotAnInteger:
        stock_in_page = paginator.page(1)
    except EmptyPage:
        stock_in_page = paginator.page(paginator.num_pages)

    return render(request, 'Inventory/stockIn/stock_in_transaction_history.html', {
        'stock_in_transactions': stock_in_page,
        'query': query,
        'page_number': page_number
    })



def handle_uploaded_stock_in_file(f):
    df = pd.read_excel(f, engine='openpyxl')
    df = df.fillna('')  # Replace NaN with an empty string

    for _, row in df.iterrows():
        # Create or update the StockInHistory record
        stock_in = StockInHistory(
            date=row.get('DATE'),
            po_number=row.get('PO NUMBER'),
            purchaser=row.get('PURCHASER'),
            item_code=row.get('ITEM CODE'),
            particulars=row.get('PARTICULAR'),
            quantity_in=row.get('QUANTITY IN'),
            unit=row.get('UNIT'),
            supplier=row.get('SUPPLIER'),
            delivery_ref=row.get('DELIVERY REF#'),
            delivery_no=row.get('DELIVERY NO.'),
            invoice_type=row.get('INVOICE TYPE'),
            invoice_no=row.get('INVOICE NO.'),
            payment_req_ref=row.get('PAYMENT REQ REF#'),
            payment_details=row.get('PAYMENT DETAILS'),
            remarks=row.get('REMARKS'),
            remarks2=row.get('REMARKS2')
        )

        # Check if the supplier folder exists or create a new one
        supplier_name = stock_in.supplier
        if supplier_name:
            folder, created = InventorySupplierFolder.objects.get_or_create(name=supplier_name)
            # Update the stock in history to reference the SupplierFolder
            stock_in.supplier_folder = folder

        # Save the stock in history record
        stock_in.save()

        # Inventory update logic:
        # Check if an inventory record for the item exists
        inventory_item = ItemInventory.objects.filter(
            item_code=stock_in.item_code,
            supplier=stock_in.supplier,
            po_product_name=stock_in.particulars,
            unit=stock_in.unit,
        ).first()

        if inventory_item:
            # Update the existing inventory record by adding the quantity_in
            inventory_item.quantity_in += stock_in.quantity_in
            inventory_item.stock = inventory_item.quantity_in - inventory_item.quantity_out
            inventory_item.save()
        else:
            # If no inventory record exists, create a new one
            ItemInventory.objects.create(
                item_code=stock_in.item_code,
                supplier=stock_in.supplier,
                po_product_name=stock_in.particulars,
                unit=stock_in.unit,
                quantity_in=stock_in.quantity_in,
                quantity_out=0,  # New stock-in, no items taken out yet
                stock=stock_in.quantity_in,  # Stock is initially the quantity in
                delivery_ref=stock_in.delivery_ref,
                delivery_no=stock_in.delivery_no,
                invoice_type=stock_in.invoice_type,
                invoice_no=stock_in.invoice_no
            )



def upload_stock_in_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                handle_uploaded_stock_in_file(request.FILES['file'])
                return JsonResponse({'status': 'success', 'message': 'Stock-in data imported successfully.'})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})
    else:
        form = UploadFileForm()
    return render(request, 'Inventory/stockIn/stock_in_upload.html', {'form': form})


