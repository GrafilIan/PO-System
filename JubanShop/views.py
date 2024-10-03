import json
import zipfile
from datetime import datetime
from decimal import Decimal
from io import BytesIO

import openpyxl
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Q, Sum
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from openpyxl.styles import PatternFill, Font
from openpyxl.workbook import Workbook

from .forms import JubanItemInventoryListForm, JubanItemInventoryQuantityForm, JubanEditRemarksForm, \
    JubanItemInventoryBulkForm, JubanStockInHistoryForm, JubanUploadFileForm
from .models import JubanItemInventory, JubanItemCodeList, JubanSiteInventoryFolder, JubanInventoryHistory, \
    JubanClientInventoryFolder, JubanCart, JubanStockInHistory, JubanInventorySupplierFolder


def juban_inventory_form(request):
    if request.method == 'POST':
        form = JubanItemInventoryListForm(request.POST)
        po_product_name = request.POST.get('po_product_name')

        # Check if the po_product_name already exists in the Juban database
        if JubanItemInventory.objects.filter(po_product_name=po_product_name).exists():
            messages.error(request, 'PO Product Name already exists. Please use a different name.')
        elif form.is_valid():
            # Save the form to create the item
            item = form.save(commit=False)  # Get the instance but don't save it to the database yet

            item.quantity_in = item.quantity_in or 0
            item.quantity_out = item.quantity_out or 0

            # Manually calculate the stock
            item.stock = item.quantity_in - item.quantity_out  # Calculate stock based on quantity_in and quantity_out

            # Now save the item with the updated stock value
            item.save()

            messages.success(request, 'Item successfully added!')
            return redirect('juban_inventory_form')  # Redirect to the same form page
        else:
            messages.error(request, 'There was an error adding the item. Please try again.')
    else:
        form = JubanItemInventoryListForm()

    return render(request, 'jubanshop/Inventory/juban_inventory_form.html', {'form': form})


def juban_inventory_table(request):
    query = request.GET.get('q')

    # Fetch all inventory items, ordered by 'po_product_name'
    inventory_items = JubanItemInventory.objects.all().order_by('po_product_name')

    # If there is a search query, filter the results
    if query:
        inventory_items = inventory_items.filter(
            Q(item_code__icontains=query) |
            Q(supplier__icontains=query) |
            Q(po_product_name__icontains=query) |
            Q(unit__icontains=query) |
            Q(quantity_in__icontains=query) |
            Q(quantity_out__icontains=query) |
            Q(stock__icontains=query)
        )

    # Ensure the ItemCodeList gets updated if necessary
    for item in inventory_items:
        if not JubanItemCodeList.objects.filter(item_code=item.item_code,
                                                po_product_name=item.po_product_name).exists():
            JubanItemCodeList.objects.create(
                item_code=item.item_code,
                po_product_name=item.po_product_name,
                unit=item.unit
            )

    # Paginate the results, displaying 100 items per page
    paginator = Paginator(inventory_items, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Render the results in the dashboard template
    return render(request, 'dashboards/juban_inventory_dashboard.html', {
        'page_obj': page_obj
    })


def juban_new_records_view(request):
    # Define the timeframe for what constitutes a "new" record (e.g., last 24 hours)
    timeframe = timezone.now() - timezone.timedelta(hours=24)

    # Get all records created in the last 24 hours, ordered by most recent
    new_records = JubanItemInventory.objects.filter(created_at__gte=timeframe).order_by('-created_at')

    return render(request, 'Jubanshop/Inventory/juban_new_records.html', {
        'new_records': new_records
    })


def juban_inventory_edit(request, id):
    inventory_item = get_object_or_404(JubanItemInventory, id=id)

    if request.method == 'POST':
        form = JubanItemInventoryQuantityForm(request.POST, instance=inventory_item)
        if form.is_valid():
            updated_item = form.save(commit=False)

            # Check if an existing record with the same supplier and po_product_name already exists
            existing_inventory_item = JubanItemInventory.objects.filter(
                supplier=updated_item.supplier,
                po_product_name=updated_item.po_product_name
            ).exclude(id=id).first()

            if existing_inventory_item:
                # If a matching record exists, update its quantities and stock
                existing_inventory_item.quantity_in += updated_item.quantity_in
                existing_inventory_item.quantity_out += updated_item.quantity_out
                existing_inventory_item.stock = existing_inventory_item.quantity_in - existing_inventory_item.quantity_out
                existing_inventory_item.save()

                # Optionally, delete the current record or return success message
                inventory_item.delete()  # Remove the current item as it merges with the existing one.
            else:
                # If no matching record, update the current inventory item
                updated_item.stock = updated_item.quantity_in - updated_item.quantity_out
                updated_item.save()

            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'status': 'error', 'errors': form.errors})
    else:
        form = JubanItemInventoryQuantityForm(instance=inventory_item)

    return render(request, 'Jubanshop/Inventory/juban_inventory_edit.html', {'form': form, 'item': inventory_item})


def juban_item_code_list(request):
    query = request.GET.get('q')  # Get the search query from the URL
    if query:
        # Filter results based on the search query
        saved_items = JubanItemCodeList.objects.filter(
            Q(item_code__icontains=query) |
            Q(po_product_name__icontains=query) |
            Q(unit__icontains=query)
        )
    else:
        saved_items = JubanItemCodeList.objects.all()  # If no search, return all items

    return render(request, 'Jubanshop/Inventory/juban_item_code_list.html',
                  {'saved_items': saved_items, 'query': query})


def juban_export_inventory_to_excel(request):
    query = request.GET.get('q')  # Search query parameter

    # Start with all inventory items
    inventory_list = JubanItemInventory.objects.all()

    # Apply search filter
    if query:
        inventory_list = inventory_list.filter(
            Q(item_code__icontains=query) |
            Q(supplier__icontains=query) |
            Q(po_product_name__icontains=query) |
            Q(new_product_name__icontains=query) |
            Q(unit__icontains=query) |
            Q(quantity_in__icontains=query) |
            Q(quantity_out__icontains=query) |
            Q(stock__icontains=query) |
            Q(price__icontains=query) |
            Q(total_amount__icontains=query)
        )

    # Create a workbook and a sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Inventory'

    header_font = Font(bold=True)
    blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    currency_format = '#,##0.00'

    # Define the headers
    headers = [
        'Item Code', 'Supplier', 'Particular', 'New Product Name', 'Unit',
        'Quantity In', 'Quantity Out', 'Stock', 'Price', 'Total Amount'
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = blue_fill

    # Populate the sheet with data
    for item in inventory_list:
        sheet.append([
            item.item_code, item.supplier, item.po_product_name, item.new_product_name,
            item.unit, item.quantity_in, item.quantity_out, item.stock,
            item.price, item.total_amount
        ])

    for cell in sheet['I']:  # Assuming price is in column I (index 9)
        if cell.row > 1:  # Skip header row
            cell.number_format = currency_format

    for cell in sheet['J']:  # Assuming total_amount is in column J (index 10)
        if cell.row > 1:  # Skip header row
            cell.number_format = currency_format

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Add extra space for better visibility
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Create an in-memory buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Set the response to return the Excel file
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Juban_Inventory.xlsx'
    return response


def juban_export_site_folder_contents(request, folder_id):
    try:
        # Get the Juban site inventory folder by ID
        folder = JubanSiteInventoryFolder.objects.get(id=folder_id)
        transactions = JubanInventoryHistory.objects.filter(site_inventory_folder=folder)

        # Create a workbook and a sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Site Inventory Records'

        # Define header styles
        header_font = Font(bold=True)
        blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        currency_format = '#,##0.00'

        # Define the headers
        headers = [
            'Transaction ID', 'Site Delivered', 'Date', 'Item', 'Quantity In', 'Quantity Out',
            'Price', 'Total Amount', 'Delivery Ref', 'Delivery No.'
        ]
        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = blue_fill
            cell.value = cell.value.upper() if cell.value is not None else cell.value

        # Populate the sheet with data
        for transaction in transactions:
            sheet.append([
                transaction.id,
                transaction.site_delivered,
                transaction.date.strftime('%Y-%m-%d') if transaction.date else 'N/A',
                str(transaction.item),  # Convert JubanItemInventory instance to string
                transaction.quantity_in,
                transaction.quantity_out,
                transaction.price,
                transaction.total_amount,
                transaction.delivery_ref,
                transaction.delivery_no,
            ])

        # Apply currency format to price and total_amount columns
        for cell in sheet['G']:  # Assuming price is in column G (index 7)
            if cell.row > 1:  # Skip header row
                cell.number_format = currency_format

        for cell in sheet['H']:  # Assuming total_amount is in column H (index 8)
            if cell.row > 1:  # Skip header row
                cell.number_format = currency_format

        # Adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Add extra space for better visibility
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Create an in-memory buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Set the response to return the Excel file
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Juban_SiteInventory_{folder.name}.xlsx'
        return response

    except JubanSiteInventoryFolder.DoesNotExist:
        return HttpResponse("Site Inventory Folder not found", status=404)


def juban_export_all_site_inventory_folders(request):
    # Create an in-memory buffer to hold the zip file
    buffer = BytesIO()

    # Create a zip file in the buffer
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Iterate through each site inventory folder
        for folder in JubanSiteInventoryFolder.objects.all():
            # Get transactions for the folder
            transactions = JubanInventoryHistory.objects.filter(site_inventory_folder=folder)

            # Create a workbook and a sheet
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'Site Inventory Records'

            # Define header styles
            header_font = Font(bold=True)
            blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
            currency_format = '#,##0.00'

            # Define the headers
            headers = [
                'Transaction ID', 'Site Delivered', 'Date', 'Item', 'Quantity In', 'Quantity Out',
                'Price', 'Total Amount', 'Delivery Ref', 'Delivery No.'
            ]
            sheet.append(headers)

            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = blue_fill
                cell.value = cell.value.upper() if cell.value is not None else cell.value

            # Populate the sheet with data
            for transaction in transactions:
                sheet.append([
                    transaction.id,
                    transaction.site_delivered,
                    transaction.date.strftime('%Y-%m-%d') if transaction.date else 'N/A',
                    str(transaction.item),  # Convert JubanItemInventory instance to string
                    transaction.quantity_in,
                    transaction.quantity_out,
                    transaction.price,
                    transaction.total_amount,
                    transaction.delivery_ref,
                    transaction.delivery_no,
                ])

            # Apply currency format to price and total_amount columns
            for cell in sheet['G']:  # Assuming price is in column G (index 7)
                if cell.row > 1:  # Skip header row
                    cell.number_format = currency_format

            for cell in sheet['H']:  # Assuming total_amount is in column H (index 8)
                if cell.row > 1:  # Skip header row
                    cell.number_format = currency_format

            # Adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook to a bytes buffer
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            # Add the workbook to the zip file
            zf.writestr(f'SiteInventoryRecords_{folder.name}.xlsx', excel_buffer.getvalue())

    buffer.seek(0)

    # Set the response to return the zip file
    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=juban_site_inventory_folders.zip'
    return response


def juban_export_client_folder_contents(request, folder_id):
    try:
        # Get the client inventory folder by ID
        folder = JubanClientInventoryFolder.objects.get(id=folder_id)
        transactions = JubanInventoryHistory.objects.filter(client_inventory_folder=folder)

        # Create a workbook and a sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Client Inventory Records'

        # Define header styles
        header_font = Font(bold=True)
        gold_fill = PatternFill(start_color='F59B00', end_color='F59B00', fill_type='solid')
        currency_format = '#,##0.00'

        # Define the headers
        headers = [
            'Transaction ID', 'Client', 'Date', 'Item', 'Quantity In', 'Quantity Out',
            'Price', 'Total Amount', 'Delivery Ref', 'Delivery No.', 'Invoice Type', 'Invoice#'
        ]
        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = gold_fill
            cell.value = cell.value.upper() if cell.value is not None else cell.value

        # Populate the sheet with data
        for transaction in transactions:
            sheet.append([
                transaction.id,
                transaction.client if transaction.client else 'N/A',
                transaction.date.strftime('%Y-%m-%d') if transaction.date else 'N/A',
                str(transaction.item),
                transaction.quantity_in,
                transaction.quantity_out,
                transaction.price,
                transaction.total_amount,
                transaction.delivery_ref,
                transaction.delivery_no,
                transaction.invoice_type,
                transaction.invoice_no,
            ])

        # Apply currency format to price and total_amount columns
        for cell in sheet['G']:
            if cell.row > 1:
                cell.number_format = currency_format

        for cell in sheet['H']:
            if cell.row > 1:
                cell.number_format = currency_format

        # Adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

        # Create an in-memory buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Set the response to return the Excel file
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=JubanClientInventory_{folder.name}.xlsx'
        return response

    except JubanClientInventoryFolder.DoesNotExist:
        return HttpResponse("Client Inventory Folder not found", status=404)


def juban_export_all_client_folders(request):
    # Create an in-memory buffer to hold the zip file
    buffer = BytesIO()

    # Create a zip file in the buffer
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Iterate through each client folder
        for folder in JubanClientInventoryFolder.objects.all():
            # Get inventory records for the folder
            transactions = JubanInventoryHistory.objects.filter(client_inventory_folder=folder)

            # Create a workbook and a sheet
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'Client Inventory Records'

            # Define header styles
            header_font = Font(bold=True)
            gold_fill = PatternFill(start_color='F59B00', end_color='F59B00', fill_type='solid')
            currency_format = '#,##0.00'

            # Define the headers
            headers = [
                'Transaction ID', 'Client', 'Date', 'Item', 'Quantity In', 'Quantity Out',
                'Price', 'Total Amount', 'Delivery Ref', 'Delivery No.', 'Invoice Type', 'Invoice#'
            ]
            sheet.append(headers)

            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = gold_fill
                cell.value = cell.value.upper() if cell.value is not None else cell.value

            # Populate the sheet with data
            for transaction in transactions:
                sheet.append([
                    transaction.id,
                    transaction.client if transaction.client else 'N/A',
                    transaction.date.strftime('%Y-%m-%d') if transaction.date else 'N/A',
                    str(transaction.item),
                    transaction.quantity_in,
                    transaction.quantity_out,
                    transaction.price,
                    transaction.total_amount,
                    transaction.delivery_ref,
                    transaction.delivery_no,
                    transaction.invoice_type,
                    transaction.invoice_no,
                ])

            # Apply currency format to price and total_amount columns
            for cell in sheet['G']:
                if cell.row > 1:
                    cell.number_format = currency_format

            for cell in sheet['H']:
                if cell.row > 1:
                    cell.number_format = currency_format

            # Adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook to a bytes buffer
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            # Add the workbook to the zip file
            zf.writestr(f'JubanClientInventory_{folder.name}.xlsx', excel_buffer.getvalue())

    buffer.seek(0)

    # Set the response to return the zip file
    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=juban_client_inventory_folders.zip'
    return response


def juban_export_transaction_history_to_excel(request):
    query = request.GET.get('q')  # Search query parameter
    date_query = request.GET.get('date')  # Date query parameter

    # Start with all transactions
    transactions = JubanInventoryHistory.objects.exclude(date__isnull=True).order_by('-date')

    # Apply search filter
    if query:
        transactions = transactions.filter(
            Q(date__icontains=query) |
            Q(item_code__icontains=query) |
            Q(supplier__icontains=query) |
            Q(po_product_name__icontains=query) |
            Q(new_product_name__icontains=query) |
            Q(unit__icontains=query) |
            Q(quantity_out__icontains=query) |
            Q(price__icontains=query) |
            Q(total_amount__icontains=query) |
            Q(site_delivered__icontains=query) |
            Q(client__icontains=query) |
            Q(delivery_ref__icontains=query) |
            Q(delivery_no__icontains=query) |
            Q(invoice_type__icontains=query) |
            Q(invoice_no__icontains=query)
        )

    # Apply date filter
    if date_query:
        try:
            date_obj = datetime.strptime(date_query, '%b %d, %Y').date()
            transactions = transactions.filter(date=date_obj)
        except ValueError:
            try:
                date_obj = datetime.strptime(date_query, '%b %Y')
                transactions = transactions.filter(date__year=date_obj.year, date__month=date_obj.month)
            except ValueError:
                try:
                    date_obj = datetime.strptime(date_query, '%B %Y')
                    transactions = transactions.filter(date__year=date_obj.year, date__month=date_obj.month)
                except ValueError:
                    try:
                        month_number = int(date_query)
                        transactions = transactions.filter(date__month=month_number)
                    except ValueError:
                        try:
                            date_obj = datetime.strptime(date_query, '%B')
                            transactions = transactions.filter(date__month=date_obj.month)
                        except ValueError:
                            pass

    # Create a workbook and a sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Juban Transaction History'

    header_font = Font(color='FFFFFF', bold=True)
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    currency_format = '#,##0.00'

    # Define the headers
    headers = [
        'Date', 'Item Code', 'Supplier', 'PO Product Name', 'New Product Name',
        'Unit', 'Quantity Out', 'Price', 'Total Amount', 'Site Delivered',
        'Client', 'Delivery Ref#', 'Delivery No.', 'Invoice Type', 'Invoice#'
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = black_fill
        cell.value = cell.value.upper() if cell.value is not None else cell.value

    # Populate the sheet with data
    for transaction in transactions:
        sheet.append([
            transaction.date.strftime('%Y-%m-%d') if transaction.date else 'N/A',
            transaction.item_code, transaction.supplier, transaction.po_product_name,
            transaction.new_product_name, transaction.unit, transaction.quantity_out,
            transaction.price, transaction.total_amount, transaction.site_delivered,
            transaction.client, transaction.delivery_ref, transaction.delivery_no,
            transaction.invoice_type, transaction.invoice_no
        ])

    # Apply currency format to Price and Total Amount columns
    for cell in sheet['H']:  # Assuming price is in column H (index 8)
        if cell.row > 1:  # Skip header row
            cell.number_format = currency_format

    for cell in sheet['I']:  # Assuming total_amount is in column I (index 9)
        if cell.row > 1:  # Skip header row
            cell.number_format = currency_format

    # Auto-size columns for better visibility
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Create an in-memory buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Set the response to return the Excel file
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=TransactionHistory.xlsx'
    return response


def juban_export_stock_in_transaction_history_to_excel(request):
    query = request.GET.get('q')  # Search query parameter
    date_query = request.GET.get('date')  # Date query parameter

    # Start with all stock-in transactions
    transactions = JubanStockInHistory.objects.exclude(date__isnull=True).order_by('-date')

    # Apply search filter
    if query:
        transactions = transactions.filter(
            Q(date__icontains=query) |
            Q(item_code__icontains=query) |
            Q(supplier__icontains=query) |
            Q(particulars__icontains=query) |
            Q(unit__icontains=query) |
            Q(quantity_in__icontains=query) |
            Q(invoice_no__icontains=query) |
            Q(invoice_type__icontains=query)
        )

    # Apply date filter
    if date_query:
        try:
            date_obj = datetime.strptime(date_query, '%b %d, %Y').date()
            transactions = transactions.filter(date=date_obj)
        except ValueError:
            try:
                date_obj = datetime.strptime(date_query, '%b %Y')
                transactions = transactions.filter(date__year=date_obj.year, date__month=date_obj.month)
            except ValueError:
                try:
                    date_obj = datetime.strptime(date_query, '%B %Y')
                    transactions = transactions.filter(date__year=date_obj.year, date__month=date_obj.month)
                except ValueError:
                    try:
                        month_number = int(date_query)
                        transactions = transactions.filter(date__month=month_number)
                    except ValueError:
                        try:
                            date_obj = datetime.strptime(date_query, '%B')
                            transactions = transactions.filter(date__month=date_obj.month)
                        except ValueError:
                            pass

    # Create a workbook and a sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Stock In Transaction History'

    header_font = Font(color='FFFFFF', bold=True)
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    # Define the headers
    headers = [
        'Date', 'PO#','Purchaser', 'Item Code', 'Particular', 'Unit',
        'Quantity In', 'Supplier', 'Remarks', 'Invoice No.', 'Invoice Type'
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = black_fill
        cell.value = cell.value.upper() if cell.value is not None else cell.value

    # Populate the sheet with data
    for transaction in transactions:
        sheet.append([
            transaction.date.strftime('%Y-%m-%d') if transaction.date else 'N/A',
            transaction.po_number, transaction.purchaser, transaction.item_code, transaction.particulars,
            transaction.unit, transaction.quantity_in, transaction.supplier,
            transaction.remarks, transaction.invoice_no, transaction.invoice_type
        ])

    # Auto-size columns for better visibility
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Create an in-memory buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Set the response to return the Excel file
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=StockInTransactionHistory.xlsx'
    return response


def juban_transaction_history(request):
    query = request.GET.get('q')  # Get search query from request
    page_number = request.GET.get('page', 1)  # Get page number from request

    # Dictionary to map full and abbreviated month names to month numbers
    month_mapping = {
        "January": "1", "February": "2", "March": "3", "April": "4",
        "May": "5", "June": "6", "July": "7", "August": "8",
        "September": "9", "October": "10", "November": "11", "December": "12",
        "Jan": "1", "Feb": "2", "Mar": "3", "Apr": "4",
        "Jun": "6", "Jul": "7", "Aug": "8", "Sep": "9", "Oct": "10", "Nov": "11", "Dec": "12"
    }

    # Retrieve all records from JubanInventoryHistory and exclude records where date is null
    transactions = JubanInventoryHistory.objects.exclude(date__isnull=True).order_by('-date')

    # Apply search filter if a query is present
    if query:
        try:
            # Try to interpret the query as a full date (e.g., 'Aug 20, 2024')
            date_obj = datetime.strptime(query, '%b %d, %Y').date()
            transactions = transactions.filter(date=date_obj)
        except ValueError:
            # If it's not a full date, check if the query is a month name (full or abbreviated)
            for month_name, month_number in month_mapping.items():
                if month_name.lower() in query.lower():
                    # Check if the query contains a year
                    year = None
                    try:
                        year = int(query.split()[-1])  # Try to extract the year
                    except (ValueError, IndexError):
                        pass

                    # Filter based on month and possibly year
                    if year:
                        transactions = transactions.filter(date__month=month_number, date__year=year)
                    else:
                        transactions = transactions.filter(date__month=month_number)
                    break
            else:
                # If not a date or month, treat as a general text search
                transactions = transactions.filter(
                    Q(item_code__icontains=query) |
                    Q(supplier__icontains=query) |
                    Q(po_product_name__icontains=query) |
                    Q(new_product_name__icontains=query) |
                    Q(unit__icontains=query) |
                    Q(quantity_out__icontains=query) |
                    Q(price__icontains=query) |
                    Q(total_amount__icontains=query) |
                    Q(site_delivered__icontains=query) |
                    Q(client__icontains=query)
                )

    # Paginate the filtered transactions
    paginator = Paginator(transactions, 100)  # Show 100 transactions per page
    try:
        transactions_page = paginator.page(page_number)
    except PageNotAnInteger:
        transactions_page = paginator.page(1)
    except EmptyPage:
        transactions_page = paginator.page(paginator.num_pages)

    return render(request, 'Jubanshop/Inventory/juban_transaction_history.html', {
        'transactions': transactions_page,
        'query': query,
        'page_number': page_number
    })


def juban_create_site_inventory_folder(request):
    if request.method == 'POST':
        folder_name = request.POST.get('folder_name')

        if folder_name:
            new_folder, created = JubanSiteInventoryFolder.objects.get_or_create(name=folder_name)

            if created:
                # Handle matching transactions here
                matching_transactions = JubanInventoryHistory.objects.filter(site_delivered=folder_name)
                for transaction in matching_transactions:
                    transaction.site_inventory_folder = new_folder
                    transaction.save()

                return JsonResponse({'success': True, 'message': 'Folder created successfully.'})
            else:
                return JsonResponse({'success': False, 'message': 'Folder with this name already exists.'})

    return JsonResponse({'success': False, 'error': 'Invalid request method.'})


def juban_delete_site_inventory_folder(request, folder_id):
    if request.method == 'POST':
        folder = get_object_or_404(JubanSiteInventoryFolder, id=folder_id)
        folder.delete()  # This will set the site_inventory_folder field in JubanInventoryHistory to NULL
        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Invalid request method.'})


def juban_site_inventory_folder_list(request):
    if request.method == 'POST':
        folder_name = request.POST.get('folder_name')

        if folder_name:
            # Create or get the folder
            new_folder, created = JubanSiteInventoryFolder.objects.get_or_create(name=folder_name)

            if created:
                return JsonResponse({'success': True, 'message': 'Folder created successfully.'})
            else:
                return JsonResponse({'success': False, 'message': 'Folder with this name already exists.'})

    # Handling GET requests to render the folder list
    folders = JubanSiteInventoryFolder.objects.all()
    context = {
        'folders': folders,
    }
    return render(request, 'Jubanshop/Inventory/juban_site_inventory_folder_list.html', context)


def juban_view_site_inventory_folder_contents(request, folder_id):
    folder = get_object_or_404(JubanSiteInventoryFolder, id=folder_id)
    transactions_list = JubanInventoryHistory.objects.filter(site_inventory_folder=folder)

    # Calculate total_amount
    total_amount = transactions_list.aggregate(total_amount_sum=Sum('total_amount'))['total_amount_sum'] or 0

    # Pagination
    paginator = Paginator(transactions_list, 100)  # Show 100 transactions per page
    page_number = request.GET.get('page')
    try:
        transactions = paginator.page(page_number)
    except PageNotAnInteger:
        transactions = paginator.page(1)
    except EmptyPage:
        transactions = paginator.page(paginator.num_pages)

    context = {
        'folder': folder,
        'transactions': transactions,
        'total_amount': total_amount,
    }

    return render(request, 'Jubanshop/Inventory/juban_site_folder_contents.html', context)


def juban_create_client_inventory_folder(request):
    if request.method == 'POST':
        folder_name = request.POST.get('folder_name')

        if folder_name:
            new_folder, created = JubanClientInventoryFolder.objects.get_or_create(name=folder_name)

            if created:
                # Handle matching transactions here
                matching_transactions = JubanInventoryHistory.objects.filter(client=folder_name)
                for transaction in matching_transactions:
                    transaction.client_inventory_folder = new_folder
                    transaction.save()

                return JsonResponse(
                    {'success': True, 'message': 'Client folder created and transactions updated successfully.'})
            else:
                return JsonResponse({'success': False, 'message': 'Client folder with this name already exists.'})

    return JsonResponse({'success': False, 'error': 'Invalid request method.'})


def juban_delete_client_inventory_folder(request, folder_id):
    if request.method == 'POST':
        folder = get_object_or_404(JubanClientInventoryFolder, id=folder_id)
        folder.delete()  # This will set the client_inventory_folder field in JubanInventoryHistory to NULL
        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Invalid request method.'})


def juban_client_inventory_folder_list(request):
    if request.method == 'POST':
        folder_name = request.POST.get('folder_name')

        if folder_name:
            # Create or get the folder
            new_folder, created = JubanClientInventoryFolder.objects.get_or_create(name=folder_name)

            if created:
                return JsonResponse({'success': True, 'message': 'Folder created successfully.'})
            else:
                return JsonResponse({'success': False, 'message': 'Folder with this name already exists.'})

    # Handling GET requests to render the folder list
    folders = JubanClientInventoryFolder.objects.all()
    context = {
        'folders': folders,
    }
    return render(request, 'Jubanshop/Inventory/juban_client_inventory_folder_list.html', context)


def juban_view_client_inventory_folder_contents(request, folder_id):
    folder = get_object_or_404(JubanClientInventoryFolder, id=folder_id)
    transactions_list = JubanInventoryHistory.objects.filter(client_inventory_folder=folder)

    total_amount = transactions_list.aggregate(total_amount_sum=Sum('total_amount'))['total_amount_sum'] or 0

    # Pagination
    paginator = Paginator(transactions_list, 100)  # Show 100 transactions per page
    page_number = request.GET.get('page')
    try:
        transactions = paginator.page(page_number)
    except PageNotAnInteger:
        transactions = paginator.page(1)
    except EmptyPage:
        transactions = paginator.page(paginator.num_pages)

    context = {
        'folder': folder,
        'transactions': transactions,
        'total_amount': total_amount,
    }

    return render(request, 'Jubanshop/Inventory/juban_client_folder_contents.html', context)


def juban_edit_inventory_history_remarks(request, folder_id, record_id):
    # Get the folder and record
    folder = get_object_or_404(JubanClientInventoryFolder, id=folder_id)
    record = get_object_or_404(JubanInventoryHistory, client_inventory_folder=folder, id=record_id)

    if request.method == 'POST':
        form = JubanEditRemarksForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            return redirect('juban_view_client_inventory_folder_contents',
                            folder_id=folder_id)  # Redirect to folder view after edit
    else:
        form = JubanEditRemarksForm(instance=record)

    context = {
        'form': form,
        'folder': folder,
        'record': record
    }
    return render(request, 'Jubanshop/Inventory/juban_client_folder_contents.html', context)


@csrf_exempt
def juban_edit_remarks(request, transaction_id):
    if request.method == 'POST':
        data = json.loads(request.body)
        remarks = data.get('remarks', '')

        try:
            transaction = JubanInventoryHistory.objects.get(id=transaction_id)
            transaction.remarks = remarks
            transaction.save()
            return JsonResponse({'success': True})
        except JubanInventoryHistory.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Transaction not found'}, status=404)

    return JsonResponse({'success': False}, status=400)


def juban_bulk_edit_inventory(request):
    if request.method == 'POST':
        if 'add_to_cart' in request.POST:
            item_ids = request.POST.getlist('item_ids')

            # Ensure there are selected items
            success = False
            for item_id in item_ids:
                quantity_out = request.POST.get(f'quantity_out_{item_id}', '0')
                quantity_out = Decimal(quantity_out) if quantity_out.strip() != '' else 0

                if quantity_out > 0:
                    success = True  # At least one valid item has been edited
                    break

            if not success:
                messages.error(request,
                               'No valid quantity out entered for any item. Please enter a quantity to add to the cart.')
                return redirect('juban_bulk_edit_inventory')

            success = True
            errors = []

            for item_id in item_ids:
                quantity_out = request.POST.get(f'quantity_out_{item_id}', '0')
                quantity_out = Decimal(quantity_out) if quantity_out.strip() != '' else 0

                if quantity_out > 0:
                    try:
                        item = JubanItemInventory.objects.get(id=item_id)
                        cart_item, created = JubanCart.objects.get_or_create(
                            item=item,
                            defaults={'quantity': quantity_out}
                        )
                        if not created:
                            cart_item.quantity += quantity_out
                            cart_item.save()

                    except JubanItemInventory.DoesNotExist:
                        success = False
                        errors.append(f"Item with ID {item_id} does not exist.")
                    except Exception as e:
                        success = False
                        errors.append(f"An error occurred for item ID {item_id}: {str(e)}")

            if success:
                messages.success(request, 'Items added to cart successfully.')
            else:
                messages.error(request, 'Some errors occurred: ' + ', '.join(errors))
            return redirect('juban_bulk_edit_inventory')

        elif 'finalize_changes' in request.POST:
            cart_items = JubanCart.objects.all()
            if not cart_items.exists():
                messages.error(request, 'No items in the cart to finalize.')
                return redirect('juban_bulk_edit_inventory')

            form = JubanItemInventoryBulkForm(request.POST)
            if form.is_valid():
                date = form.cleaned_data['date']
                location_type = form.cleaned_data['location_type']
                location_name = form.cleaned_data['location_name']
                delivery_ref = form.cleaned_data['delivery_ref']
                delivery_no = form.cleaned_data['delivery_no']
                invoice_type = form.cleaned_data['invoice_type']
                invoice_no = form.cleaned_data['invoice_no']

                success = True
                errors = []

                for cart_item in cart_items:
                    try:
                        item = cart_item.item
                        quantity_out = cart_item.quantity
                        price = item.price
                        total_amount = quantity_out * price

                        if location_type == 'site':
                            folder, created = JubanSiteInventoryFolder.objects.get_or_create(name=location_name)
                            item.site_inventory_folder = folder
                            item.site_delivered = location_name

                        elif location_type == 'client':
                            client_folder, created = JubanClientInventoryFolder.objects.get_or_create(
                                name=location_name)
                            item.client_inventory_folder = client_folder
                            item.client = location_name

                        item.date = date
                        item.quantity_out += quantity_out
                        item.stock = item.quantity_in - item.quantity_out
                        item.total_amount = total_amount
                        item.delivery_ref = delivery_ref
                        item.delivery_no = delivery_no
                        item.invoice_type = invoice_type
                        item.invoice_no = invoice_no
                        item.site_or_client_choice = location_type
                        item.save()

                        JubanInventoryHistory.objects.create(
                            item=item,
                            date=date,
                            item_code=item.item_code,
                            supplier=item.supplier,
                            po_product_name=item.po_product_name,
                            unit=item.unit,
                            quantity_in=item.quantity_in,
                            quantity_out=quantity_out,
                            stock=item.stock,
                            price=price,
                            total_amount=total_amount,
                            delivery_ref=delivery_ref,
                            delivery_no=delivery_no,
                            invoice_type=invoice_type,
                            invoice_no=invoice_no,
                            site_inventory_folder=item.site_inventory_folder if location_type == 'site' else None,
                            client_inventory_folder=item.client_inventory_folder if location_type == 'client' else None,
                            site_delivered=item.site_delivered if location_type == 'site' else None,
                            client=item.client if location_type == 'client' else None,
                        )

                        cart_item.delete()

                    except JubanItemInventory.DoesNotExist:
                        success = False
                        errors.append(f"Item with ID {cart_item.item.id} does not exist.")
                    except Exception as e:
                        success = False
                        errors.append(f"An error occurred for item ID {cart_item.item.id}: {str(e)}")

                JubanCart.objects.all().delete()

                if success:
                    messages.success(request, 'Items updated successfully.')
                else:
                    messages.error(request, 'Some errors occurred: ' + ', '.join(errors))
                return redirect('juban_bulk_edit_inventory')

    else:
        query = request.GET.get('q', '')
        if query:
            items = JubanItemInventory.objects.filter(po_product_name__icontains=query)
        else:
            items = JubanItemInventory.objects.all()

        cart_items = JubanCart.objects.all()
        for cart_item in cart_items:
            cart_item.total_amount = cart_item.quantity * cart_item.item.price

        form = JubanItemInventoryBulkForm()

        return render(request, 'Jubanshop/Inventory/juban_bulk_edit_inventory.html', {
            'form': form,
            'items': items,
            'cart_items': cart_items,
            'query': query,
        })


def juban_remove_cart_item(request, cart_item_id):
    cart_item = get_object_or_404(JubanCart, id=cart_item_id)
    cart_item.delete()
    messages.success(request, 'Item removed from cart successfully.')
    return redirect('juban_bulk_edit_inventory')


def juban_stock_in_create(request):
    if request.method == 'POST':
        form = JubanStockInHistoryForm(request.POST)
        if form.is_valid():
            # Save the new stock in history but don't commit to DB yet
            stock_in = form.save(commit=False)

            # Check if the particulars match an existing po_product_name in JubanItemInventory
            if not JubanItemInventory.objects.filter(po_product_name=stock_in.particulars).exists():
                # Raise an error if the item does not exist in the JubanItemInventory list
                return JsonResponse({'status': 'error',
                                     'message': 'The item does not exist in the item list. Please check the particulars.'})

            # Get the supplier name from the stock in history
            supplier_name = stock_in.supplier

            if supplier_name:
                # Create or get the JubanInventorySupplierFolder associated with this supplier
                folder, created = JubanInventorySupplierFolder.objects.get_or_create(name=supplier_name)

                # Associate the new stock in history with the supplier's folder
                stock_in.supplier_folder = folder
                stock_in.save()

                # Ensure all existing stock in histories with the same supplier are associated with the folder
                matching_histories = JubanStockInHistory.objects.filter(supplier=supplier_name,
                                                                        supplier_folder__isnull=True)
                for history in matching_histories:
                    history.supplier_folder = folder
                    history.save()

            # Update or create the related JubanItemInventory record
            item_inventory, created = JubanItemInventory.objects.get_or_create(
                item_code=stock_in.item_code,
                supplier=stock_in.supplier,
                defaults={
                    'po_product_name': stock_in.particulars,
                    'unit': stock_in.unit,
                    'quantity_in': stock_in.quantity_in,
                    'quantity_out': 0,  # Set quantity_out to 0 if the record is new
                    'stock': stock_in.quantity_in  # Initialize stock to quantity_in for new items
                }
            )

            if not created:
                # Update the existing JubanItemInventory record by adding the quantity_in
                item_inventory.quantity_in += stock_in.quantity_in
                item_inventory.stock = item_inventory.quantity_in - item_inventory.quantity_out
                item_inventory.quantity_out = item_inventory.quantity_out or 0  # Ensure quantity_out is never None
                item_inventory.save()

            return JsonResponse({'status': 'success'})
        else:
            print(form.errors)  # Debugging form errors
            return JsonResponse({'status': 'error', 'errors': form.errors})
    else:
        form = JubanStockInHistoryForm()

    return render(request, 'Jubanshop/stockIn/juban_stock_in_form.html', {'form': form})


def juban_inventory_supplier_list_folders(request):
    if request.method == 'POST':
        folder_name = request.POST.get('folder_name')

        if folder_name:
            # Create or get the JubanInventorySupplierFolder based on the folder_name
            new_folder, created = JubanInventorySupplierFolder.objects.get_or_create(name=folder_name)

            if created:
                # Handle matching stock in records here
                matching_stock_ins = JubanStockInHistory.objects.filter(supplier=folder_name)
                for stock_in in matching_stock_ins:
                    stock_in.supplier_folder = new_folder
                    stock_in.save()

                return JsonResponse({'success': True, 'message': 'Folder created successfully.'})
            else:
                return JsonResponse({'success': False, 'message': 'Folder with this name already exists.'})

    # Handling GET requests to render the folder list
    folders = JubanInventorySupplierFolder.objects.all()
    context = {
        'folders': folders,
    }
    return render(request, 'Jubanshop/stockIn/juban_inventory_supplier_list_folders.html', context)


def juban_inventory_supplier_contents(request, folder_id):
    folder = get_object_or_404(JubanInventorySupplierFolder, id=folder_id)

    # Get the search query from the GET parameters
    query = request.GET.get('q', '').strip()
    page_number = request.GET.get('page', 1)

    # Initial queryset
    stock_in_histories = JubanStockInHistory.objects.filter(supplier_folder=folder)

    # If there's a search query, filter the stock in records accordingly
    if query:
        try:
            # Try to interpret the query as a month number (e.g., "08" for August)
            month_number = int(query)
            stock_in_histories = stock_in_histories.filter(date__month=month_number)
        except ValueError:
            try:
                # Try to interpret the query as a full month name only (e.g., "August")
                month_mapping = {
                    "January": "1", "February": "2", "March": "3", "April": "4",
                    "May": "5", "June": "6", "July": "7", "August": "8",
                    "September": "9", "October": "10", "November": "11", "December": "12"
                }
                month_number = month_mapping.get(query.capitalize())
                if month_number:
                    stock_in_histories = stock_in_histories.filter(date__month=month_number)
                else:
                    raise ValueError("Not a valid month name")
            except ValueError:
                # If not a date, treat it as a general text search
                stock_in_histories = stock_in_histories.filter(
                    Q(po_number__icontains=query) |
                    Q(purchaser__icontains=query) |
                    Q(item_code__icontains=query) |
                    Q(particulars__icontains=query) |
                    Q(quantity_in__icontains=query) |
                    Q(unit__icontains=query) |
                    Q(fbbd_ref_number__icontains=query) |
                    Q(remarks__icontains=query) |
                    Q(supplier__icontains=query) |
                    Q(delivery_ref__icontains=query) |
                    Q(delivery_no__icontains=query) |
                    Q(invoice_type__icontains=query) |
                    Q(invoice_no__icontains=query) |
                    Q(payment_req_ref__icontains=query) |
                    Q(payment_details__icontains=query) |
                    Q(remarks2__icontains=query)
                )

    # Pagination
    paginator = Paginator(stock_in_histories, 100)  # Show 100 records per page
    try:
        stock_in_records = paginator.page(page_number)
    except PageNotAnInteger:
        stock_in_records = paginator.page(1)
    except EmptyPage:
        stock_in_records = paginator.page(paginator.num_pages)

    # Render the template with the context
    context = {
        'folder': folder,
        'stock_in_records': stock_in_records,
        'query': query,
        'page_number': page_number,
    }
    return render(request, 'Jubanshop/stockIn/juban_inventory_supplier_contents.html', context)


def juban_export_inventory_supplier_contents(request, folder_id):
    try:
        # Get the JubanInventorySupplierFolder by ID
        folder = JubanInventorySupplierFolder.objects.get(id=folder_id)
        stock_in_list = JubanStockInHistory.objects.filter(supplier_folder=folder)

        # Create a workbook and a sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Stock In Records'

        # Define header styles
        header_font = Font(bold=True)
        blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        currency_format = '#,##0.00'

        # Define the headers
        headers = [
            'Date', 'PO Number', 'Purchaser', 'Item Code', 'Particular',
            'Quantity In', 'Unit', 'FBBD Ref#', 'Remarks', 'Supplier',
            'Delivery Ref#', 'Delivery No.', 'Invoice Type', 'Invoice No.',
            'Payment Req Ref#', 'Payment Details', 'Remarks2'
        ]
        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = blue_fill
            cell.value = cell.value.upper() if cell.value is not None else cell.value

        # Populate the sheet with data
        for stock_in in stock_in_list:
            sheet.append([
                stock_in.date.strftime('%Y-%m-%d') if stock_in.date else 'N/A',
                stock_in.po_number,
                stock_in.purchaser,
                stock_in.item_code,
                stock_in.particulars,
                stock_in.quantity_in,
                stock_in.unit,
                stock_in.fbbd_ref_number,
                stock_in.remarks,
                stock_in.supplier,
                stock_in.delivery_ref,
                stock_in.delivery_no,
                stock_in.invoice_type,
                stock_in.invoice_no,
                stock_in.payment_req_ref,
                stock_in.payment_details,
                stock_in.remarks2
            ])

        for cell in sheet['F']:  # Assuming quantity_in is in column F (index 6)
            if cell.row > 1:  # Skip header row
                cell.number_format = currency_format

        # Auto-adjust column widths based on content
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Add extra space for better visibility
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Create an in-memory buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Set the response to return the Excel file
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=StockInRecords_{folder.name}.xlsx'
        return response

    except JubanInventorySupplierFolder.DoesNotExist:
        return HttpResponse("Juban Inventory Supplier Folder not found", status=404)


def juban_get_item_details(request):
    particulars = request.GET.get('particulars', None)

    if particulars:
        try:
            item = JubanItemInventory.objects.get(po_product_name=particulars)
            return JsonResponse({
                'item_code': item.item_code,
                'unit': item.unit,
                'supplier': item.supplier,
            })
        except JubanItemInventory.DoesNotExist:
            return JsonResponse({'error': 'Item not found'}, status=404)

    return JsonResponse({'error': 'Invalid request'}, status=400)


def juban_get_item_inventory(request):
    items = list(JubanItemInventory.objects.values('id', 'po_product_name'))
    return JsonResponse({
        'items': [{'id': item['id'], 'text': item['po_product_name']} for item in items]
    })


def juban_stock_in_transaction_history(request):
    query = request.GET.get('q')  # Get search query from request
    page_number = request.GET.get('page', 1)  # Get page number from request

    # Dictionary to map full and abbreviated month names to month numbers
    month_mapping = {
        "January": "1", "February": "2", "March": "3", "April": "4",
        "May": "5", "June": "6", "July": "7", "August": "8",
        "September": "9", "October": "10", "November": "11", "December": "12",
        "Jan": "1", "Feb": "2", "Mar": "3", "Apr": "4",
        "Jun": "6", "Jul": "7", "Aug": "8", "Sep": "9", "Oct": "10", "Nov": "11", "Dec": "12"
    }

    # Retrieve all records from JubanStockInHistory and exclude records where date is null
    stock_in_transactions = JubanStockInHistory.objects.exclude(date__isnull=True).order_by('-date')

    # Apply search filter if a query is present
    if query:
        try:
            # Try to interpret the query as a full date (e.g., 'Aug 20, 2024')
            date_obj = datetime.strptime(query, '%b %d, %Y').date()
            stock_in_transactions = stock_in_transactions.filter(date=date_obj)
        except ValueError:
            # If it's not a full date, check if the query is a month name (full or abbreviated)
            for month_name, month_number in month_mapping.items():
                if month_name.lower() in query.lower():
                    year = None
                    try:
                        year = int(query.split()[-1])  # Try to extract the year
                    except (ValueError, IndexError):
                        pass

                    if year:
                        stock_in_transactions = stock_in_transactions.filter(date__month=month_number, date__year=year)
                    else:
                        stock_in_transactions = stock_in_transactions.filter(date__month=month_number)
                    break
            else:
                # If not a date or month, treat as a general text search
                stock_in_transactions = stock_in_transactions.filter(
                    Q(item_code__icontains=query) |
                    Q(supplier__icontains=query) |
                    Q(particulars__icontains=query) |
                    Q(unit__icontains=query) |
                    Q(quantity_in__icontains=query) |
                    Q(remarks__icontains=query) |
                    Q(invoice_no__icontains=query) |
                    Q(po_number__icontains=query)
                )

    # Paginate the filtered stock-in transactions
    paginator = Paginator(stock_in_transactions, 100)  # Show 100 transactions per page
    try:
        stock_in_page = paginator.page(page_number)
    except PageNotAnInteger:
        stock_in_page = paginator.page(1)
    except EmptyPage:
        stock_in_page = paginator.page(paginator.num_pages)

    return render(request, 'Jubanshop/stockIn/juban_stock_in_transaction_history.html', {
        'stock_in_transactions': stock_in_page,
        'query': query,
        'page_number': page_number
    })


def juban_handle_uploaded_stock_in_file(f):
    df = pd.read_excel(f, engine='openpyxl')
    df = df.fillna('')  # Replace NaN with an empty string

    for _, row in df.iterrows():
        # Create or update the JubanStockInHistory record
        stock_in = JubanStockInHistory(
            date=row.get('DATE'),
            po_number=row.get('PO NUMBER'),
            purchaser=row.get('PURCHASER'),
            item_code=row.get('ITEM CODE'),
            particulars=row.get('PARTICULAR'),
            quantity_in=row.get('QUANTITY IN'),
            unit=row.get('UNIT'),
            supplier=row.get('SUPPLIER'),
            delivery_ref=row.get('DELIVERY REF#'),
            delivery_no=row.get('DELIVERY NO.'),
            invoice_type=row.get('INVOICE TYPE'),
            invoice_no=row.get('INVOICE NO.'),
            payment_req_ref=row.get('PAYMENT REQ REF#'),
            payment_details=row.get('PAYMENT DETAILS'),
            remarks=row.get('REMARKS'),
            remarks2=row.get('REMARKS2')
        )

        # Check if the supplier folder exists or create a new one
        supplier_name = stock_in.supplier
        if supplier_name:
            folder, created = JubanInventorySupplierFolder.objects.get_or_create(name=supplier_name)
            # Update the stock in history to reference the SupplierFolder
            stock_in.supplier_folder = folder

        # Save the stock in history record
        stock_in.save()

        # Inventory update logic:
        # Check if an inventory record for the item exists
        inventory_item = JubanItemInventory.objects.filter(
            item_code=stock_in.item_code,
            supplier=stock_in.supplier,
            po_product_name=stock_in.particulars,
            unit=stock_in.unit,
        ).first()

        if inventory_item:
            # Update the existing inventory record by adding the quantity_in
            inventory_item.quantity_in += stock_in.quantity_in
            inventory_item.stock = inventory_item.quantity_in - inventory_item.quantity_out
            inventory_item.save()
        else:
            # If no inventory record exists, create a new one
            JubanItemInventory.objects.create(
                item_code=stock_in.item_code,
                supplier=stock_in.supplier,
                po_product_name=stock_in.particulars,
                unit=stock_in.unit,
                quantity_in=stock_in.quantity_in,
                quantity_out=0,  # New stock-in, no items taken out yet
                stock=stock_in.quantity_in,  # Stock is initially the quantity in
                delivery_ref=stock_in.delivery_ref,
                delivery_no=stock_in.delivery_no,
                invoice_type=stock_in.invoice_type,
                invoice_no=stock_in.invoice_no
            )


def juban_upload_stock_in_file(request):
    if request.method == 'POST':
        form = JubanUploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                juban_handle_uploaded_stock_in_file(request.FILES['file'])
                return JsonResponse({'status': 'success', 'message': 'Stock-in data imported successfully.'})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})
    else:
        form = JubanUploadFileForm()

    return render(request, 'Jubanshop/stockIn/juban_stock_in_upload.html', {'form': form})
